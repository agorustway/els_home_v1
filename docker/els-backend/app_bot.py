# [v5.0.42] DNS 장애 환경 대응: 전역 소켓 패치 적용 (가장 먼저 실행)
import dns_fix
dns_fix.apply_dns_patch()

import os
import sys
import json
import threading
import subprocess
import time
import uuid
import io
import math
from pathlib import Path
from datetime import datetime, timedelta, timezone
try:
    from flask import Flask, request, jsonify, Response, send_file
    from flask_cors import CORS
except ModuleNotFoundError:
    class Flask:
        def __init__(self, *args, **kwargs):
            self.logger = type("Logger", (), {"warning": lambda *a, **k: None})()
        def route(self, *args, **kwargs):
            def deco(fn):
                return fn
            return deco
        def run(self, *args, **kwargs):
            return None
    request = None
    def jsonify(*args, **kwargs):
        return args[0] if args else kwargs
    class Response:
        def __init__(self, *args, **kwargs):
            pass
    send_file = None
    def CORS(*args, **kwargs):
        return None
from urllib.request import Request, urlopen
try:
    import pandas as pd
except ModuleNotFoundError:
    pd = None

# --- KST 설정 ---
KST = timezone(timedelta(hours=9))

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

ELSBOT_DIR = Path("/app/elsbot")
if not ELSBOT_DIR.exists(): ELSBOT_DIR = Path("elsbot").resolve()
DAEMON_URL = os.environ.get("DAEMON_URL", "http://127.0.0.1:31999")
LAST_RESULT_FILE = ELSBOT_DIR / "last_search_result.json"

file_store = {}
global_progress = {"total": 0, "completed": 0, "is_running": False}
# [v4.9.8] 유휴 감지용: 마지막으로 개별 컨테이너 조회가 완료된 시각
global_last_activity_time = 0
global_stop_requested = threading.Event()

def _status_row(cn, code, message):
    return [str(cn or "").strip().upper(), code, message] + [""] * 12

def _should_retry_rows(rows):
    if not rows:
        return True
    first = rows[0] if isinstance(rows[0], (list, tuple)) else []
    code = str(first[1] if len(first) > 1 else "")
    message = str(first[2] if len(first) > 2 else "")
    if code != "ERROR":
        return False
    if "WORKER_RETIRED" in message or "WORKER_NOT_READY" in message:
        return True
    non_retryable = [
        "유효하지 않은 컨테이너 번호",
        "비밀번호",
        "로그인 3회",
        "보안 모드",
        "조회 중지됨",
        "INPUT_NOT_FOUND",
        "메뉴 진입 실패",
        "워커 대기 시간 초과",
    ]
    if any(token in message for token in non_retryable):
        return False
    retryable = [
        "WORKER_RETIRED",
        "WORKER_NOT_READY",
        "이전 조회 결과 잔상 감지",
        "데이터 추출 실패",
        "timed out",
        "timeout",
    ]
    return any(token in message for token in retryable)

def _pop_ready_ordered_results(pending_results, next_index):
    ready = []
    while next_index in pending_results:
        ready.append(pending_results.pop(next_index))
        next_index += 1
    return ready, next_index

def _daemon_health(timeout=3):
    try:
        r = urlopen(Request(DAEMON_URL + "/health", method="GET"), timeout=timeout)
        if r.getcode() == 200:
            return json.loads(r.read().decode("utf-8"))
    except Exception:
        pass
    return {}

def _configured_batch_workers(configured_workers=None):
    try:
        raw = os.environ.get("ELS_BATCH_MAX_WORKERS", 4) if configured_workers is None else configured_workers
        configured = int(raw)
    except (TypeError, ValueError):
        configured = 4
    return max(1, configured)

def _effective_batch_workers(health, configured_workers=None, reserve_single=True, in_flight=0):
    """실제로 비어 있거나 현재 배치가 점유 중인 워커 수만큼 배치 병렬도를 잡는다."""
    configured = _configured_batch_workers(configured_workers)
    total = int(health.get("total_drivers") or 0)
    if total <= 0:
        return 0
    try:
        available = int(health.get("available_drivers"))
    except (TypeError, ValueError):
        available = total
    try:
        occupied_by_this_batch = max(0, int(in_flight or 0))
    except (TypeError, ValueError):
        occupied_by_this_batch = 0

    max_batch_by_total = total
    if reserve_single and total > 1:
        max_batch_by_total = total - 1
    live_window = max(0, available) + occupied_by_this_batch
    if live_window <= 0:
        return 0
    return max(1, min(configured, max_batch_by_total, live_window))

def _daemon_login_once(user_id, user_pw, show_browser=False, timeout=380):
    body = json.dumps({
        "userId": user_id,
        "userPw": user_pw,
        "showBrowser": show_browser,
    }, ensure_ascii=False).encode("utf-8")
    req = Request(DAEMON_URL + "/login", data=body, method="POST", headers={"Content-Type": "application/json"})
    r = urlopen(req, timeout=timeout)
    return json.loads(r.read().decode("utf-8"))

def _daemon_available():
    try:
        r = urlopen(Request(DAEMON_URL + "/health", method="GET"), timeout=2)
        return r.getcode() == 200
    except Exception: return False

@app.route("/health", methods=["GET"])
def health(): return jsonify({"status": "ok", "service": "els-bot"})

@app.route("/api/els/capabilities", methods=["GET"])
def capabilities():
    global global_progress, global_last_activity_time
    daemon_status = {
        "available": False,
        "driver_active": False,
        "user_id": None,
        "workers": [],
        "max_drivers": 4,
        "total_drivers": 0,
        "available_drivers": 0,
        "is_logging_in": False,
        "active_init_threads": 0,
    }
    data = _daemon_health(timeout=3)
    if data:
        daemon_status["available"] = True
        daemon_status["driver_active"] = data.get("driver_active", False)
        daemon_status["user_id"] = data.get("user_id")
        daemon_status["workers"] = data.get("workers", [])
        daemon_status["max_drivers"] = data.get("max_drivers", 4)
        daemon_status["total_drivers"] = data.get("total_drivers", 0)
        daemon_status["available_drivers"] = data.get("available_drivers", 0)
        daemon_status["is_logging_in"] = data.get("is_logging_in", False)
        daemon_status["active_init_threads"] = data.get("active_init_threads", 0)

    # [v4.9.8] 좀비 잠금 해제: 전체 작업 5분 초과 시 강제 종료
    if global_progress.get("is_running") and global_progress.get("start_time"):
        elapsed = time.time() - global_progress["start_time"]
        if elapsed > 300:  # 5분 초과
            app.logger.warning(f"[좀비복구] 작업이 {int(elapsed)}초 경과하여 강제 종료 처리")
            global_progress["is_running"] = False
            global_progress["completed"] = global_progress["total"]

    # [v4.9.8] 유휴 잠금 해제: 마지막 개별 조회 완료 후 3분간 추가 조회 없으면 종료로 간주
    if global_progress.get("is_running") and global_last_activity_time > 0:
        idle = time.time() - global_last_activity_time
        if idle > 180:  # 3분 무활동
            app.logger.warning(f"[유휴복구] 마지막 조회 후 {int(idle)}초 무활동. 잠금 해제")
            global_progress["is_running"] = False
            global_progress["completed"] = global_progress["total"]

    return jsonify({
        "available": daemon_status["available"],
        "driver_active": daemon_status["driver_active"],
        "user_id": daemon_status["user_id"],
        "progress": global_progress,
        "workers": daemon_status["workers"],
        "max_drivers": daemon_status["max_drivers"],
        "total_drivers": daemon_status["total_drivers"],
        "available_drivers": daemon_status["available_drivers"],
        "is_logging_in": daemon_status["is_logging_in"],
        "active_init_threads": daemon_status["active_init_threads"],
        "batch_workers": _effective_batch_workers(daemon_status),
        "parseAvailable": True
    })

@app.route("/api/els/logs", methods=["GET"])
def logs():
    try:
        r = urlopen(Request(DAEMON_URL + "/logs", method="GET"), timeout=3)
        if r.getcode() == 200:
            return Response(r.read(), mimetype="application/json")
    except Exception: pass
    return jsonify({"ok": False, "log": []})

@app.route("/api/els/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    if not _daemon_available(): return jsonify({"ok": False, "error": "ELS 데몬 응답 없음"}), 404

    def generate():
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        req = Request(DAEMON_URL + "/login", data=body, method="POST", headers={"Content-Type": "application/json"})
        login_result = {}
        def _thread():
            try:
                r = urlopen(req, timeout=400)
                login_result['resp'] = json.loads(r.read().decode("utf-8"))
            except Exception as e: login_result['error'] = str(e)
        t = threading.Thread(target=_thread)
        t.start()
        sent_logs = set()
        while t.is_alive():
            try:
                l_req = urlopen(DAEMON_URL + "/logs", timeout=1)
                l_data = json.loads(l_req.read().decode("utf-8"))
                for line in l_data.get("log", []):
                    if line not in sent_logs:
                        yield f"LOG:{line}\n"
                        sent_logs.add(line)
            except Exception: pass
            time.sleep(1)
        t.join()
        
        # [Fix] 최종 로그 스윕
        try:
            l_req = urlopen(DAEMON_URL + "/logs", timeout=1)
            l_data = json.loads(l_req.read().decode("utf-8"))
            for line in l_data.get("log", []):
                if line not in sent_logs:
                    yield f"LOG:{line}\n"; sent_logs.add(line)
        except Exception: pass
        res = login_result.get('resp', {"ok": False, "error": login_result.get('error', 'Unknown')})
        yield "RESULT:" + json.dumps(res, ensure_ascii=False) + "\n"

    return Response(generate(), mimetype="text/plain; charset=utf-8")

@app.route("/api/els/run", methods=["POST"])
def run():
    data = request.get_json(silent=True) or {}
    containers = data.get("containers", [])
    uid = data.get("userId", "")
    pw = data.get("userPw", "")
    show_browser = data.get("showBrowser", False)

    global global_progress, global_last_activity_time
    global_stop_requested.clear()
    global_progress = {"total": len(containers), "completed": 0, "is_running": True, "start_time": time.time()}
    global_last_activity_time = time.time()

    def generate():
        global global_progress, global_last_activity_time
        try:
            final_rows = []
            headers = ["컨테이너번호", "No", "수출입", "구분", "터미널", "MOVE TIME", "모선", "항차", "선사", "적공", "SIZE", "POD", "POL", "차량번호", "RFID"]
            from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
            
            def fetch(index, cn):
                last_error = None
                for attempt in range(1, 3):
                    if global_stop_requested.is_set():
                        return index, {"ok": True, "result": [_status_row(cn, "ERROR", "조회 중지됨")]}, cn, attempt, "사용자 중지"
                    try:
                        body = json.dumps({
                            "userId": uid,
                            "userPw": pw,
                            "containerNo": cn,
                            "showBrowser": show_browser,
                            "requestPurpose": "batch",
                            "acquireTimeoutSec": 45,
                        }, ensure_ascii=False).encode("utf-8")
                        req = Request(DAEMON_URL + "/run", data=body, method="POST", headers={"Content-Type": "application/json"})
                        r = urlopen(req, timeout=300)
                        data = json.loads(r.read().decode("utf-8"))
                        rows = data.get("result") or []
                        if _should_retry_rows(rows) and attempt < 2:
                            last_error = rows[0][2] if rows and len(rows[0]) > 2 else data.get("error", "불확실 실패")
                            time.sleep(1.2)
                            continue
                        return index, data, cn, attempt, last_error
                    except Exception as e:
                        last_error = str(e)
                        if attempt < 2:
                            time.sleep(1.2)
                            continue
                        return index, {"ok": False, "error": str(e), "result": [_status_row(cn, "ERROR", str(e))]}, cn, attempt, last_error

            sent_logs = set()
            reserve_single = data.get("reserveSingle")
            if reserve_single is None:
                reserve_single = os.environ.get("ELS_RESERVE_SINGLE_WORKER", "false").lower() == "true"
            else:
                reserve_single = bool(reserve_single)
            daemon_health = _daemon_health(timeout=3)
            if int(daemon_health.get("total_drivers") or 0) <= 0 and (not uid or not pw):
                for cn in containers:
                    rows = [_status_row(cn, "ERROR", "LOGIN_REQUIRED: saved login is not ready")]
                    final_rows.extend(rows)
                    global_progress["completed"] += 1
                    yield "RESULT_PARTIAL:" + json.dumps({"result": rows}, ensure_ascii=False) + "\n"
                yield "RESULT:" + json.dumps({"ok": True, "result": final_rows}, ensure_ascii=False) + "\n"
                return
            if int(daemon_health.get("total_drivers") or 0) <= 0 and uid and pw:
                yield "LOG:[session] no ready worker; starting daemon login before batch\n"
                try:
                    login_res = _daemon_login_once(uid, pw, show_browser=show_browser)
                    for line in login_res.get("log", []):
                        yield f"LOG:{line}\n"
                    if not login_res.get("ok"):
                        message = login_res.get("error") or login_res.get("message") or "login failed"
                        for cn in containers:
                            rows = [_status_row(cn, "ERROR", f"LOGIN_READY_FAILED: {message}")]
                            final_rows.extend(rows)
                            global_progress["completed"] += 1
                            yield "RESULT_PARTIAL:" + json.dumps({"result": rows}, ensure_ascii=False) + "\n"
                        yield "RESULT:" + json.dumps({"ok": True, "result": final_rows}, ensure_ascii=False) + "\n"
                        return
                    daemon_health = _daemon_health(timeout=3)
                except Exception as e:
                    for cn in containers:
                        rows = [_status_row(cn, "ERROR", f"LOGIN_READY_FAILED: {e}")]
                        final_rows.extend(rows)
                        global_progress["completed"] += 1
                        yield "RESULT_PARTIAL:" + json.dumps({"result": rows}, ensure_ascii=False) + "\n"
                    yield "RESULT:" + json.dumps({"ok": True, "result": final_rows}, ensure_ascii=False) + "\n"
                    return
            configured_workers = _configured_batch_workers()
            batch_workers = _effective_batch_workers(daemon_health, configured_workers=configured_workers, reserve_single=reserve_single)
            max_executor_workers = max(configured_workers, batch_workers)
            ready_wait_timeout = float(os.environ.get("ELS_BATCH_READY_WAIT_SEC", 90))
            yield f"LOG:⚙️ 배치 병렬도 {batch_workers}개 적용 (가용 {daemon_health.get('available_drivers', 0)}/활성 {daemon_health.get('total_drivers', 0)}/{daemon_health.get('max_drivers', 4)}, 단건/AI 예약 {'ON' if reserve_single else 'OFF'})\n"
            executor = ThreadPoolExecutor(max_workers=max_executor_workers)
            try:
                futures = {}
                completed_results = {}
                emitted_indices = set()
                next_submit_index = 0
                last_worker_refresh = 0
                ready_wait_started = time.time()
                last_ready_wait_log = 0

                def refresh_batch_workers(force=False):
                    nonlocal batch_workers, daemon_health, last_worker_refresh
                    now = time.time()
                    if not force and now - last_worker_refresh < 3:
                        return None
                    last_worker_refresh = now
                    latest = _daemon_health(timeout=1)
                    if not latest:
                        return None
                    daemon_health = latest
                    new_workers = _effective_batch_workers(
                        latest,
                        configured_workers=configured_workers,
                        reserve_single=reserve_single,
                        in_flight=len(futures),
                    )
                    if new_workers != batch_workers:
                        old_workers = batch_workers
                        batch_workers = new_workers
                        return old_workers, new_workers, latest
                    return None

                def mark_stopped_from(start_index=0):
                    for idx in range(start_index, len(containers)):
                        if idx not in completed_results:
                            completed_results[idx] = {
                                "cn": containers[idx],
                                "rows": [_status_row(containers[idx], "ERROR", "조회 중지됨")],
                                "attempts": 1,
                                "retry_reason": "사용자 중지",
                                "error": "조회 중지됨",
                            }

                def submit_more():
                    nonlocal next_submit_index
                    while (
                        next_submit_index < len(containers)
                        and len(futures) < batch_workers
                        and not global_stop_requested.is_set()
                    ):
                        if 0 < next_submit_index < batch_workers:
                            time.sleep(1)
                        idx = next_submit_index
                        futures[executor.submit(fetch, idx, containers[idx])] = idx
                        next_submit_index += 1

                submit_more()

                while futures or next_submit_index < len(containers):
                    if not futures and next_submit_index < len(containers):
                        worker_change = refresh_batch_workers(force=True)
                        if worker_change:
                            old_workers, new_workers, latest = worker_change
                            yield f"LOG:[batch] worker window {old_workers}->{new_workers} (가용 {latest.get('available_drivers', 0)}/활성 {latest.get('total_drivers', 0)}/{latest.get('max_drivers', 4)})\n"
                        submit_more()
                        if not futures:
                            now = time.time()
                            if now - ready_wait_started > ready_wait_timeout:
                                for idx in range(next_submit_index, len(containers)):
                                    completed_results[idx] = {
                                        "cn": containers[idx],
                                        "rows": [_status_row(containers[idx], "ERROR", "워커 준비 대기 시간 초과")],
                                        "attempts": 1,
                                        "retry_reason": "워커 준비 대기 시간 초과",
                                        "error": "워커 준비 대기 시간 초과",
                                    }
                                break
                            if now - last_ready_wait_log >= 10:
                                last_ready_wait_log = now
                                yield f"LOG:가용 워커가 없어 남은 조회를 잠시 대기합니다. (가용 {daemon_health.get('available_drivers', 0)}/활성 {daemon_health.get('total_drivers', 0)})\n"
                            try:
                                l_req = urlopen(DAEMON_URL + "/logs", timeout=1)
                                l_data = json.loads(l_req.read().decode("utf-8"))
                                for line in l_data.get("log", []):
                                    if line not in sent_logs:
                                        yield f"LOG:{line}\n"
                                        sent_logs.add(line)
                            except Exception:
                                pass
                            time.sleep(0.5)
                            continue
                    else:
                        ready_wait_started = time.time()

                    if global_stop_requested.is_set():
                        for f, idx in list(futures.items()):
                            if f.cancel():
                                completed_results[idx] = {
                                    "cn": containers[idx],
                                    "rows": [_status_row(containers[idx], "ERROR", "조회 중지됨")],
                                    "attempts": 1,
                                    "retry_reason": "사용자 중지",
                                    "error": "조회 중지됨",
                                }
                                del futures[f]
                        mark_stopped_from(next_submit_index)
                        yield "LOG:⏹️ 조회 중지 요청을 받아 남은 작업을 취소합니다.\n"
                        if not futures:
                            break

                    # 1. 태스크 완료 확인 및 결과 배출
                    done, _ = wait(futures.keys(), timeout=0.5, return_when=FIRST_COMPLETED)
                    for f in done:
                        try:
                            index, res, cn, attempts, retry_reason = f.result()
                        except Exception as e:
                            index = futures.get(f, 0)
                            cn = containers[index] if index < len(containers) else ""
                            attempts = 1
                            retry_reason = str(e)
                            res = {"ok": False, "error": str(e), "result": [_status_row(cn, "ERROR", str(e))]}
                        rows = res.get("result") or [_status_row(cn, "ERROR", res.get("error") or "조회 실패")]
                        completed_results[index] = {
                            "cn": cn,
                            "rows": rows,
                            "attempts": attempts,
                            "retry_reason": retry_reason,
                            "error": res.get("error"),
                        }
                        del futures[f]
                        item = completed_results[index]
                        rows = item["rows"]
                        emitted_indices.add(index)
                        global_progress["completed"] += 1
                        global_last_activity_time = time.time()  # [v4.9.8] 개별 조회 완료 시마다 갱신
                        if item["attempts"] > 1:
                            yield f"LOG:↻ [{item['cn']}] 1회 재조회 후 확정 ({item['retry_reason']})\n"
                        yield "RESULT_PARTIAL:" + json.dumps({"result": rows}, ensure_ascii=False) + "\n"

                    worker_change = refresh_batch_workers()
                    if worker_change:
                        old_workers, new_workers, latest = worker_change
                        yield f"LOG:[batch] worker window {old_workers}->{new_workers} (가용 {latest.get('available_drivers', 0)}/활성 {latest.get('total_drivers', 0)}/{latest.get('max_drivers', 4)})\n"
                    submit_more()

                    # 2. 로그 실시간 스트리밍 (루프 도중)
                    try:
                        l_req = urlopen(DAEMON_URL + "/logs", timeout=1)
                        l_data = json.loads(l_req.read().decode("utf-8"))
                        for line in l_data.get("log", []):
                            if line not in sent_logs:
                                yield f"LOG:{line}\n"
                                sent_logs.add(line)
                    except Exception: pass

                for index in sorted(completed_results):
                    if index in emitted_indices:
                        continue
                    item = completed_results[index]
                    rows = item["rows"]
                    emitted_indices.add(index)
                    global_progress["completed"] += 1
                    global_last_activity_time = time.time()
                    yield "RESULT_PARTIAL:" + json.dumps({"result": rows}, ensure_ascii=False) + "\n"

                for index in range(len(containers)):
                    item = completed_results.get(index)
                    if item:
                        final_rows.extend(item["rows"])
            finally:
                executor.shutdown(wait=False, cancel_futures=True)

            # 3. [Fix] 모든 작업 완료 후 남아있는 로그 최종 스윕 (릴레이 지연 고려)
            time.sleep(1)
            try:
                l_req = urlopen(DAEMON_URL + "/logs", timeout=2)
                l_data = json.loads(l_req.read().decode("utf-8"))
                for line in l_data.get("log", []):
                    if line not in sent_logs:
                        yield f"LOG:{line}\n"
                        sent_logs.add(line)
            except Exception: pass

            # 엑셀 생성 (openpyxl - 2시트, 서식, 틀고정)
            if final_rows:
                token = str(uuid.uuid4())[:8]
                from collections import defaultdict
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                grouped = defaultdict(list)
                for row in final_rows:
                    grouped[str(row[0])].append(row)

                no1_rows, all_sorted = [], []
                for cn, rows_g in grouped.items():
                    sorted_r = sorted(rows_g, key=lambda r: int(r[1]) if str(r[1]).isdigit() else 999)
                    all_sorted.extend(sorted_r)
                    no1 = next((r for r in sorted_r if str(r[1]) == '1'), None)
                    if no1:
                        no1_rows.append(no1)

                wb = openpyxl.Workbook()
                h_font   = Font(name='맑은 고딕', size=10, bold=True)
                d_font   = Font(name='맑은 고딕', size=10)
                imp_font = Font(name='맑은 고딕', size=10, color='B91C1C')
                inb_font = Font(name='맑은 고딕', size=10, color='1D4ED8')
                h_fill   = PatternFill('solid', fgColor='F2F2F2')
                imp_fill = PatternFill('solid', fgColor='FEE2E2')
                inb_fill = PatternFill('solid', fgColor='EFF6FF')
                h_align  = Alignment(horizontal='center', vertical='center')
                d_align  = Alignment(vertical='center')
                th_side  = Side(style='thin', color='94A3B8')
                td_side  = Side(style='thin', color='E2E8F0')
                h_border = Border(top=th_side, left=th_side, bottom=th_side, right=th_side)
                d_border = Border(top=td_side, left=td_side, bottom=td_side, right=td_side)

                def write_ws(ws, data_rows):
                    ws.append(headers)
                    for r in data_rows:
                        ws.append([str(v) if v is not None else '' for v in r])
                    for row_cells in ws.iter_rows():
                        for cell in row_cells:
                            if cell.row == 1:
                                cell.font = h_font; cell.fill = h_fill
                                cell.alignment = h_align; cell.border = h_border
                            else:
                                val = str(cell.value or '')
                                cell.alignment = d_align; cell.border = d_border
                                if '수입' in val:
                                    cell.fill = imp_fill; cell.font = imp_font
                                elif '반입' in val:
                                    cell.fill = inb_fill; cell.font = inb_font
                                else:
                                    cell.font = d_font
                    ws.freeze_panes = 'A2'
                    if ws.dimensions:
                        ws.auto_filter.ref = ws.dimensions
                    for col_cells in ws.columns:
                        max_len = 0
                        for cell in col_cells:
                            l = sum(2 if ord(c) > 127 else 1 for c in str(cell.value or ''))
                            max_len = max(max_len, l)
                        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 50)
                    ws.row_dimensions[1].height = 18

                ws1 = wb.active; ws1.title = '최신이력_No1'
                write_ws(ws1, no1_rows)
                ws2 = wb.create_sheet('전체이력')
                write_ws(ws2, all_sorted)

                out = io.BytesIO()
                wb.save(out); out.seek(0)
                now_kst = datetime.now(KST)
                file_name = f"컨테이너이력조회_{now_kst.strftime('%Y%m%d%H%M%S')}.xlsx"
                file_store[token] = {'data': out.read(), 'name': file_name}
                yield "RESULT:" + json.dumps({"ok": True, "result": final_rows, "downloadToken": token, "fileName": file_name}, ensure_ascii=False) + "\n"
        except GeneratorExit:
            global_stop_requested.set()
            try:
                req = Request(DAEMON_URL + "/stop", data=b"{}", method="POST", headers={"Content-Type": "application/json"})
                urlopen(req, timeout=2)
            except Exception:
                pass
            raise
        finally:
            # [v4.9.8] 핵심: 스트림 중단(스마트폰 창 내림 등)에도 반드시 잠금 해제
            global_progress["is_running"] = False
            global_progress["completed"] = global_progress.get("total", 0)

    return Response(generate(), mimetype="text/plain; charset=utf-8")
        
# --- [v5.6.2] AI 어시스턴트 전용: 단일 컨테이너 실시간 조회 (JSON 반환) ---
@app.route("/api/els/container/tracking", methods=["GET"])
def container_tracking():
    cntr_no = request.args.get("cntrNo")
    if not cntr_no: return jsonify({"ok": False, "error": "컨테이너 번호 누락"}), 400
    
    # 봇 데몬에게 단일 조회 요청
    try:
        body = json.dumps({
            "containerNo": cntr_no,
            "requestPurpose": "single",
            "acquireTimeoutSec": 240,
        }, ensure_ascii=False).encode("utf-8")
        req = Request(DAEMON_URL + "/run", data=body, method="POST", headers={"Content-Type": "application/json"})
        r = urlopen(req, timeout=300)
        resp_data = json.loads(r.read().decode("utf-8"))
        
        if not resp_data.get("ok"):
            return jsonify({"ok": False, "error": resp_data.get("error", "조회 실패")})
            
        # 결과 파싱 (성공 또는 내역없음)
        rows = resp_data.get("result", [])
        tracking_list = []
        for row in rows:
            if len(row) >= 15:
                tracking_list.append({
                    "no": row[1], "type": row[2], "status": row[3], "terminal": row[4],
                    "time": row[5], "vessel": row[6], "voyage": row[7], "line": row[8],
                    "full_empty": row[9], "size": row[10], "pod": row[11], "pol": row[12],
                    "vehicle": row[13], "rfid": row[14]
                })
        
        return jsonify({"ok": True, "tracking_list": tracking_list})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/els/download/<token>", methods=["GET"])
def download(token):
    item = file_store.pop(token, None)
    if not item: return "Expired", 404
    buf  = item['data'] if isinstance(item, dict) else item
    name = item.get('name', 'Container_History.xlsx') if isinstance(item, dict) else (request.args.get('filename') or 'Container_History.xlsx')
    from urllib.parse import quote
    safe = quote(name, safe='')
    return Response(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe}"})

@app.route("/api/els/stop-daemon", methods=["POST"])
def stop_daemon():
    """[v4.9.8] 데몬 세션 강제 종료 + 잠금 해제"""
    global global_progress, global_last_activity_time
    global_stop_requested.set()
    if _daemon_available():
        try:
            req = Request(DAEMON_URL + "/stop", method="POST")
            urlopen(req, timeout=5)
        except Exception: pass
    # 백엔드 잠금도 동시 해제
    global_progress["is_running"] = False
    global_progress["completed"] = global_progress.get("total", 0)
    global_last_activity_time = 0
    return jsonify({"ok": True, "message": "데몬 세션 및 잠금 해제 완료"})

@app.route("/api/els/screenshot", methods=["GET"])
def screenshot():
    idx = request.args.get('idx', '1')
    try:
        r = urlopen(DAEMON_URL + f"/screenshot?idx={idx}", timeout=5)
        return Response(r.read(), mimetype='image/png')
    except Exception: return "Screenshot fail", 404

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=2931, threaded=True)
