# [v5.0.42] DNS 장애 환경 대응: 전역 소켓 패치 적용 (가장 먼저 실행)
import dns_fix
dns_fix.apply_dns_patch()

import os
import sys
import json
import threading
import subprocess
import tempfile
import uuid
import logging
import time
import re
import io
import math
import hashlib
import gc
import requests
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

import pandas as pd
from werkzeug.exceptions import HTTPException
from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS 
from werkzeug.utils import secure_filename
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, timezone
from supabase import create_client, Client
from file_sync_gate import StableFileSyncGate
from asan_performance import register_asan_performance_routes

# --- KST 설정 ---
KST = timezone(timedelta(hours=9))

# --- Supabase 설정 ---
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

# --- 로깅 설정 ---
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] [%(levelname)s] In %(module)s: %(message)s')

app = Flask(__name__)
# --- [수정] CORS 설정: 모든 로컬 포트 + 외부 도메인 허용 ---
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20MB

ELSBOT_DIR = Path("/app/elsbot")
# Local fallback if /app/elsbot doesn't exist (running on host)
if not ELSBOT_DIR.exists():
    ELSBOT_DIR = Path("elsbot").resolve()

RUNNER = ELSBOT_DIR / "els_web_runner.py"
CONFIG_PATH = ELSBOT_DIR / "els_config.json"
# [수정] 데몬 주소를 환경변수에서 가져오도록 변경
DAEMON_URL = os.environ.get("DAEMON_URL", "http://127.0.0.1:31999")

# --- 전역 변수 ---
file_store = {}
# [추가] 진행률 트래킹용 전역 변수
global_progress = {"total": 0, "completed": 0, "is_running": False}
# [v4.9.8] 유휴 감지용: 마지막으로 개별 컨테이너 조회가 완료된 시각
global_last_activity_time = 0
LAST_RESULT_FILE = ELSBOT_DIR / "last_search_result.json"

def _env_int(name, default, minimum=0):
    try:
        return max(minimum, int(os.environ.get(name, default)))
    except (TypeError, ValueError):
        return default

ASAN_DISPATCH_SYNC_POLL_SECONDS = _env_int("ASAN_DISPATCH_SYNC_POLL_SECONDS", 60, 15)
ASAN_DISPATCH_SYNC_QUIET_SECONDS = _env_int("ASAN_DISPATCH_SYNC_QUIET_SECONDS", 8, 0)
ASAN_DISPATCH_SYNC_RETRY_SECONDS = _env_int("ASAN_DISPATCH_SYNC_RETRY_SECONDS", 60, 10)
ASAN_SHIPPING_SYNC_POLL_SECONDS = _env_int("ASAN_SHIPPING_SYNC_POLL_SECONDS", 60, 30)
ASAN_SHIPPING_SYNC_QUIET_SECONDS = _env_int("ASAN_SHIPPING_SYNC_QUIET_SECONDS", 8, 0)
ASAN_SHIPPING_SYNC_RETRY_SECONDS = _env_int("ASAN_SHIPPING_SYNC_RETRY_SECONDS", 90, 10)
ASAN_DISPATCH_SETTINGS_CACHE_SECONDS = _env_int("ASAN_DISPATCH_SETTINGS_CACHE_SECONDS", 300, 30)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_HOUR = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_HOUR", 3, 0)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MINUTE = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MINUTE", 10, 0)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_FAIL_LIMIT = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_FAIL_LIMIT", 10, 1)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MAX_TARGETS = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MAX_TARGETS", 10000, 1)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_TIMEOUT_SECONDS = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_TIMEOUT_SECONDS", 3600, 300)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_ACQUIRE_TIMEOUT_SECONDS = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_ACQUIRE_TIMEOUT_SECONDS", 300, 30)
ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_SUBMIT_DELAY_SECONDS = _env_int("ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_SUBMIT_DELAY_SECONDS", 2, 0)
ELS_BOT_API_URL = os.environ.get("ELS_BOT_API_URL", "http://127.0.0.1:2931").rstrip("/")

dispatch_sync_gate = StableFileSyncGate(
    quiet_seconds=ASAN_DISPATCH_SYNC_QUIET_SECONDS,
    retry_seconds=ASAN_DISPATCH_SYNC_RETRY_SECONDS,
)
shipping_sync_gate = StableFileSyncGate(
    quiet_seconds=ASAN_SHIPPING_SYNC_QUIET_SECONDS,
    retry_seconds=ASAN_SHIPPING_SYNC_RETRY_SECONDS,
)

register_asan_performance_routes(app, supabase, KST)

# --- 전역 에러 핸들러 ---
@app.errorhandler(Exception)
def handle_global_exception(e):
    # [수정] 404 등 일반적인 HTTP 에러는 조용히 처리함 (로그 도배 방지)
    if isinstance(e, HTTPException):
        return jsonify({
            "ok": False,
            "error": str(e.description),
            "code": e.code
        }), e.code
        
    app.logger.error("An unhandled exception occurred: %s", str(e), exc_info=True)
    response = {
        "ok": False,
        "error": "An unexpected server error occurred.",
        "log": [f"[FATAL] {type(e).__name__}: {str(e)}"]
    }
    return jsonify(response), 500

# --- 활동 로그 (Activity Logs) ---
@app.route("/api/logs", methods=["POST"])
def post_log():
    """Vercel의 /api/logs를 대체하는 로그 저장 API"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        data = request.get_json(silent=True) or {}
        # Next.js의 로그 수집 형식 그대로 수용
        metadata = data.get("metadata", {})
        
        # [신규] 클라이언트 IP 수집 (anonymous 접근 등 식별용)
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if client_ip:
            # 쉼표로 구분된 다중 IP일 경우 첫 번째(원래 클라이언트) IP 선택
            metadata["ip"] = client_ip.split(',')[0].strip()
            
        # user_email이 anonymous일 경우 IP를 식별자로 병기
        user_email = data.get("user_email") or data.get("email", "anonymous")
        if user_email == "anonymous" and client_ip:
            user_email = f"anonymous ({metadata['ip']})"
            
        log_entry = {
            "user_id": metadata.get("user_id"), # metadata에서 id 추출
            "user_email": user_email,
            "action_type": data.get("action_type") or data.get("type", "PAGE_VIEW"),
            "path": data.get("path", "/"),
            "metadata": metadata
        }
        res = supabase.from_("user_activity_logs").insert(log_entry).execute()
        return jsonify({"ok": True})
    except Exception as e:
        app.logger.error(f"로그 저장 실패: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/debug/log", methods=["POST"])
def post_debug_log():
    """안드로이드 등 외부 앱 전용 디버그 로그 수신 (debug_app.log)"""
    try:
        data = request.get_json(silent=True) or {}
        ts = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
        with open("debug_app.log", "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {json.dumps(data, ensure_ascii=False)}\n")
        return jsonify({"ok": True})
    except Exception as e:
        app.logger.error(f"디버그 로그 저장 실패: {e}")
        return jsonify({"error": str(e)}), 500

# --- [v4.4.40] 아산지점 배차판 자동 동기화 로직 ---
last_mtime_cache = {}
dispatch_settings_cache = {"data": None, "loaded_at": 0.0}


def get_asan_dispatch_settings(force=False):
    now_ts = time.time()
    cached = dispatch_settings_cache.get("data")
    if (
        not force
        and cached
        and now_ts - float(dispatch_settings_cache.get("loaded_at") or 0) < ASAN_DISPATCH_SETTINGS_CACHE_SECONDS
    ):
        return cached

    try:
        res = supabase.from_("branch_dispatch_settings").select("*").eq("branch_id", "asan").single().execute()
        settings = res.data
        if settings:
            dispatch_settings_cache["data"] = settings
            dispatch_settings_cache["loaded_at"] = now_ts
        return settings
    except Exception as exc:
        if cached:
            app.logger.warning(f"[자동동기화] 배차 설정 조회 실패, 캐시 사용: {exc}")
            return cached
        raise


def _dispatch_db_has_current_mtime(dtype, mtime_ts, mtime):
    """컨테이너 재시작 직후 캐시가 비어도 DB가 최신이면 대용량 엑셀 파싱을 건너뛴다."""
    if not supabase:
        return False
    try:
        res = (
            supabase.from_("branch_dispatch")
            .select("file_modified_at")
            .eq("branch_id", "asan")
            .eq("type", dtype)
            .order("file_modified_at", desc=True)
            .limit(1)
            .execute()
        )
        db_mtime = (res.data or [{}])[0].get("file_modified_at")
        if not db_mtime:
            return False
        db_ts = datetime.fromisoformat(str(db_mtime).replace("Z", "+00:00")).timestamp()
        if abs(db_ts - mtime_ts) < 1:
            last_mtime_cache[dtype] = mtime
            return True
    except Exception as exc:
        app.logger.warning(f"[자동동기화] {dtype} DB 파일수정일 확인 실패: {exc}")
    return False


def sync_asan_dispatch_python(force=False):
    """나스 엑셀 파일을 읽어 Supabase를 업데이트하는 Python 버전 로직"""
    global last_mtime_cache
    if not supabase: return
    try:
        if force:
            app.logger.info("[자동동기화] 아산 배차판 수동 동기화 시작...")
        settings = get_asan_dispatch_settings(force=force)
        if not settings:
            app.logger.error("[자동동기화] 설정을 찾을 수 없습니다.")
            return

        for dtype in ['glovis', 'mobis']:
            rel_path = settings.get(f"{dtype}_path")
            if not rel_path: continue
            
            full_path = Path("/app/data") / rel_path.lstrip("/")
            
            if not full_path.exists():
                app.logger.warning(f"[자동동기화] 파일을 찾을 수 없음: {full_path} (rel_path: {rel_path})")
                continue
            
            # 파일 수정 시간
            file_stat = full_path.stat()
            mtime_ts = file_stat.st_mtime
            mtime = datetime.fromtimestamp(mtime_ts, tz=KST).isoformat()
            cached_mtime = last_mtime_cache.get(dtype)
            file_signature = (mtime_ts, file_stat.st_size)
            
            # 변경 감지 (force 옵션이 없으면 캐시 확인)
            if not force and cached_mtime == mtime:
                continue

            if not force and not cached_mtime and _dispatch_db_has_current_mtime(dtype, mtime_ts, mtime):
                dispatch_sync_gate.mark_synced(f"dispatch:{dtype}", file_signature)
                app.logger.info(f"[자동동기화] {dtype} DB 최신 상태 확인, 컨테이너 재시작 후 최초 전체 파싱 생략")
                continue

            if cached_mtime:
                app.logger.info(f"[자동동기화] {dtype} 파일 변경 감지: 이전={cached_mtime}, 현재={mtime}")
            else:
                app.logger.info(f"[자동동기화] {dtype} 최초 동기화 확인: 파일수정일={mtime}")

            decision = dispatch_sync_gate.check(f"dispatch:{dtype}", file_signature, force=force)
            if not decision.ready:
                if decision.reason in ("pending", "settling"):
                    app.logger.info(f"[자동동기화] {dtype} 파일 저장 안정화 대기 중 ({decision.reason})")
                continue
            
            app.logger.info(f"[자동동기화] 파일 변경 확인됨. 데이터 추출 시작... ({dtype})")
            
            # 네트워크 파일 행(Hang) 방지를 위해 로컬 임시 파일로 복사
            import tempfile, shutil
            temp_path = tempfile.mktemp(suffix=".xlsx")
            
            try:
                shutil.copy2(full_path, temp_path)
                app.logger.info(f"[자동동기화] {dtype} 로컬 임시 파일 복사 완료: {temp_path}")
            except Exception as e:
                app.logger.error(f"[자동동기화] {dtype} 로컬 임시 파일 복사 실패: {e}")
                continue
            
            # 엑셀 읽기
            xl = None
            wb = None
            df = None
            data_df = None
            rows = None
            comments_dict = None
            xl = pd.ExcelFile(temp_path)
            sync_count = 0
            
            # 숫자 형태의 시트 이름 필터링 (최신 날짜 우선)
            now = datetime.now(KST)
            current_month = now.month
            date_sheets = []
            
            all_sheets = xl.sheet_names
            app.logger.info(f"[자동동기화] {dtype} 파일 전체 시트 목록: {all_sheets}")
            
            for s in all_sheets:
                match = re.search(r'(\d+)[\./](\d+)', s)
                if match:
                    m = int(match.group(1))
                    d = int(match.group(2))
                    
                    # 연도 롤오버를 고려한 정렬용 가중치 계산
                    # 현재가 1, 2월인데 시트가 11, 12월이면 작년 자료로 취급 (가중치 낮춤)
                    # 현재가 11, 12월인데 시트가 1, 2월이면 내년 자료로 취급 (가중치 높임)
                    sort_score = m * 100 + d
                    if current_month <= 3 and m >= 10:
                        sort_score -= 1200 # 작년
                    elif current_month >= 10 and m <= 3:
                        sort_score += 1200 # 내년
                    
                    date_sheets.append((s, m, d, sort_score))
                    app.logger.info(f"[자동동기화] 날짜 시트 발견: {s} -> ({m}월 {d}일, 점수: {sort_score})")
            
            if not date_sheets:
                app.logger.warning(f"[자동동기화] {dtype} 파일에서 날짜 형식의 시트를 하나도 찾지 못했습니다.")
                continue
            
            # 가중치 점수 순으로 정렬 (최신순으로 처리하되, 모든 시트 동기화)
            # 초절전 메모리 최적화: read_only=True 모드 사용
            wb = None
            try:
                import openpyxl
                import gc
                wb = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)
            except Exception as e:
                app.logger.warning(f"[자동동기화] openpyxl 로드 실패: {e}")

            for sheet_name, month, day, sort_score in date_sheets:
                try:
                    app.logger.info(f"[자동동기화] {dtype} 시트 처리 중: {sheet_name} ({month}/{day})")

                    # 타겟 날짜 계산 (연도 결정)
                    year = now.year
                    if current_month <= 3 and month >= 10:
                        year -= 1 # 작년 11~12월
                    elif current_month >= 10 and month <= 3:
                        year += 1 # 내년 1~2월
                    target_date = f"{year}-{month:02d}-{day:02d}"

                    # 해당 날짜/타입의 기존 데이터 삭제 - UPSERT 도입으로 스킵 (v5.5.14)
                    # supabase.from_("branch_dispatch").delete().eq("branch_id", "asan").eq("type", dtype).eq("target_date", target_date).execute()

                    # 시트 파싱
                    df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                    header_idx = -1
                    for i, row in df.head(100).iterrows():
                        if row.astype(str).str.contains('구분').any():
                            header_idx = i
                            break
                    
                    if header_idx < 0:
                        app.logger.warning(f"[자동동기화] '{sheet_name}' 시트 건너뜀: 헤더 미발견")
                        continue
                    
                    headers = df.iloc[header_idx].fillna('').astype(str).map(lambda x: x.replace('\n', ' ').strip()).tolist()
                    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]
                    data_df = df.iloc[header_idx + 1:]
                    
                    filter_col = 12 if dtype == 'glovis' else 15
                    rows = []
                    comments_dict = {}
                    row_idx_in_db = 0
                    
                    # 메모 추출 (read_only 모드에서도 가능)
                    sheet_comments = {}
                    if wb and sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        header_col_idx = -1
                        for j, h in enumerate(headers):
                            if '구분' in str(h):
                                header_col_idx = j
                                break
                        if header_col_idx == -1: header_col_idx = 0
                            
                        openpyxl_r_offset = 0
                        openpyxl_c_offset = 0
                        found_header = False
                        # read_only 모드에서는 iter_rows가 가장 효율적
                        for r_idx, r_cells in enumerate(ws.iter_rows(min_row=1, max_row=100)):
                            for c_idx, cell in enumerate(r_cells):
                                val = str(cell.value) if cell.value else ""
                                if '구분' in val:
                                    openpyxl_r_offset = r_idx - header_idx
                                    openpyxl_c_offset = c_idx - header_col_idx
                                    found_header = True
                                    break
                            if found_header: break
                        
                        if found_header:
                            for r_idx, r_cells in enumerate(ws.iter_rows()):
                                for c_idx, cell in enumerate(r_cells):
                                    if hasattr(cell, 'comment') and cell.comment:
                                        pd_r = r_idx - openpyxl_r_offset
                                        pd_c = c_idx - openpyxl_c_offset
                                        sheet_comments[(pd_r, pd_c)] = cell.comment.text
                    
                    orig_index_list = data_df.index.tolist()
                    for i_pos, orig_iloc_idx in enumerate(orig_index_list):
                        row = data_df.loc[orig_iloc_idx]
                        # '합계' 포함 시 건너뜀 (break 대신 continue로 변경하여 데이터 유실 방지)
                        if any(str(c).find('합계') >= 0 for c in row if pd.notnull(c)):
                            continue
                        f_val = str(row.iloc[filter_col]) if filter_col < len(row) else ''
                        if not f_val or f_val == '0' or f_val == 'nan':
                            continue
                        row_list = row.fillna('').astype(str).tolist()
                        rows.append(row_list)
                        for c_idx in range(len(row_list)):
                            cmt = sheet_comments.get((orig_iloc_idx, c_idx))
                            if cmt: comments_dict[f"{row_idx_in_db}:{c_idx}"] = str(cmt)
                        row_idx_in_db += 1
                    
                    if rows: 
                        # upsert로 변경하여 중복 데이터 방지
                        supabase.from_("branch_dispatch").upsert({
                            "branch_id": "asan",
                            "type": dtype,
                            "target_date": target_date,
                            "headers": headers,
                            "data": rows,
                            "comments": comments_dict,
                            "file_modified_at": mtime,
                            "updated_at": now.isoformat()
                        }, on_conflict="branch_id,type,target_date").execute()
                        sync_count += 1
                        app.logger.info(f"[자동동기화] {dtype} - {target_date} 완료 ({len(rows)}건)")
                    
                    # 매 시트 처리 후 메모리 강제 해제
                    del sheet_comments
                    df = None
                    data_df = None
                    rows = None
                    comments_dict = None
                    gc.collect()

                except Exception as sheet_err:
                    app.logger.error(f"[자동동기화] 시트 '{sheet_name}' 처리 중 오류: {sheet_err}")

            if wb: wb.close()
            gc.collect()

            last_mtime_cache[dtype] = mtime
            dispatch_sync_gate.mark_synced(f"dispatch:{dtype}", file_signature)
            
            # 임시 파일 삭제
            try:
                import os
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception as e:
                app.logger.warning(f"[자동동기화] {dtype} 임시 파일 삭제 실패: {e}")
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            try:
                if xl:
                    xl.close()
            except Exception:
                pass
            xl = None
            wb = None
            df = None
            data_df = None
            rows = None
            comments_dict = None
            gc.collect()
                
    except Exception as e:
        app.logger.error(f"[자동동기화] 치명적 오류: {e}")

def asan_sync_scheduler():
    """배경에서 시간을 체크하여 동기화를 수행하는 스레드"""
    app.logger.info(
        f"[스케줄러] 아산 배차판 자동 동기화 스케줄러 시작 "
        f"(poll={ASAN_DISPATCH_SYNC_POLL_SECONDS}s, quiet={ASAN_DISPATCH_SYNC_QUIET_SECONDS}s)"
    )
    
    while True:
        try:
            now = datetime.now(KST)
            # 06:00 ~ 23:00 사이 (주말 포함 매일)
            if 6 <= now.hour <= 23:
                # 매 루프마다 수정 여부를 체크하고, 수정된 경우만 동기화
                sync_asan_dispatch_python()
            
            time.sleep(ASAN_DISPATCH_SYNC_POLL_SECONDS)
        except Exception as e:
            app.logger.error(f"[스케줄러] 루프 오류: {e}")
            time.sleep(ASAN_DISPATCH_SYNC_POLL_SECONDS)

# 스케줄러 시작
threading.Thread(target=asan_sync_scheduler, daemon=True).start()

@app.route("/api/branches/asan/sync", methods=["POST"])
def manual_asan_sync():
    """아산지점 배차판 수동 동기화 (프론트엔드 버튼 클릭 시)"""
    try:
        app.logger.info("[수동동기화] 프론트엔드에서 아산 배차판 동기화 요청 (force=True)")
        sync_asan_dispatch_python(force=True)
        return jsonify({
            "results": [
                {"type": "glovis", "success": True, "sheets": "backend-sync"},
                {"type": "mobis", "success": True, "sheets": "backend-sync"}
            ]
        })
    except Exception as e:
        app.logger.error(f"[수동동기화] 오류: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/logs", methods=["GET"])
def get_logs():
    """활동 로그 조회 (날짜 범위 검색 포함)"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        email = request.args.get("email", "")
        log_type = request.args.get("type", "")
        start_date = request.args.get("startDate", "")
        end_date = request.args.get("endDate", "")
        page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 30))

        query = supabase.from_("user_activity_logs").select("*", count="exact")
        if email: query = query.ilike("user_email", f"%{email}%")
        if log_type: query = query.eq("action_type", log_type)
        if start_date: query = query.gte("created_at", f"{start_date} 00:00:00")
        if end_date: query = query.lte("created_at", f"{end_date} 23:59:59")
        
        query = query.order("created_at", desc=True)
        
        # 페이징
        start = (page - 1) * limit
        end = start + limit - 1
        query = query.range(start, end)
        
        res = query.execute()
        
        return jsonify({
            "logs": res.data,
            "pagination": {
                "page": page,
                "limit": limit,
                "total": res.count,
                "totalPages": math.ceil((res.count or 0) / limit)
            }
        })
    except Exception as e:
        app.logger.error(f"로그 조회 실패: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/logs", methods=["DELETE"])
def delete_logs():
    """로그 삭제 (Vercel API와 동일 사양)"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        data = request.get_json(silent=True) or {}
        delete_type = data.get("deleteType") # 'ALL' or 'DATE'
        date_before = data.get("dateBefore")
        
        query = supabase.from_("user_activity_logs").delete()
        
        if delete_type == "ALL":
            # 모든 데이터 삭제 (id가 null이 아닌 것)
            query = query.neq("id", "00000000-0000-0000-0000-000000000000")
        elif date_before:
            query = query.lt("created_at", date_before)
        else:
            return jsonify({"error": "Invalid deletion criteria"}), 400
            
        res = query.execute()
        return jsonify({"success": True})
    except Exception as e:
        app.logger.error(f"로그 삭제 실패: {e}")
        return jsonify({"error": str(e)}), 500

def _vehicle_haversine_km(lat1, lng1, lat2, lng2):
    p = math.pi / 180
    a = (
        0.5
        - math.cos((lat2 - lat1) * p) / 2
        + math.cos(lat1 * p) * math.cos(lat2 * p) * (1 - math.cos((lng2 - lng1) * p)) / 2
    )
    return 12742 * math.asin(math.sqrt(max(0, a)))


def _vehicle_point_ms(point):
    raw = (point or {}).get("recorded_at") or (point or {}).get("timestamp") or (point or {}).get("created_at")
    if not raw:
        return 0
    try:
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00")).timestamp() * 1000
    except Exception:
        return 0


def _vehicle_filter_locations(points):
    ordered = []
    for p in points or []:
        try:
            lat = float(p.get("lat"))
            lng = float(p.get("lng"))
        except Exception:
            continue
        if not (33 <= lat <= 39.5 and 124 <= lng <= 132):
            continue
        item = dict(p)
        item["lat"] = lat
        item["lng"] = lng
        try:
            item["speed"] = float(item.get("speed") or 0)
        except Exception:
            item["speed"] = 0
        try:
            item["accuracy"] = float(item.get("accuracy") or 0)
        except Exception:
            item["accuracy"] = 0
        ordered.append(item)

    ordered.sort(key=_vehicle_point_ms)
    filtered = []
    for idx, curr in enumerate(ordered):
        if curr.get("accuracy", 0) > 120:
            continue
        prev = filtered[-1] if filtered else None
        if not prev:
            filtered.append(curr)
            continue

        time_sec = max(1, (_vehicle_point_ms(curr) - _vehicle_point_ms(prev)) / 1000)
        if _vehicle_point_ms(curr) + 1000 < _vehicle_point_ms(prev):
            continue
        dist_km = _vehicle_haversine_km(prev["lat"], prev["lng"], curr["lat"], curr["lng"])
        implied = dist_km / (time_sec / 3600)
        sensor = max(0, curr.get("speed", 0))
        speed_limit = 60 if sensor <= 4 else (90 if sensor < 15 else min(145, max(105, sensor + 45)))
        if dist_km > 0.05 and implied > speed_limit:
            continue
        if sensor <= 4 and dist_km > 0.06 and implied > 25:
            continue
        if sensor < 15 and dist_km > 0.08 and implied > 45:
            continue

        nxt = ordered[idx + 1] if idx + 1 < len(ordered) else None
        if nxt:
            next_dist = _vehicle_haversine_km(curr["lat"], curr["lng"], float(nxt["lat"]), float(nxt["lng"]))
            bridge_dist = _vehicle_haversine_km(prev["lat"], prev["lng"], float(nxt["lat"]), float(nxt["lng"]))
            if sensor < 15 and dist_km > 0.08 and next_dist > 0.08 and bridge_dist < max(0.06, dist_km * 0.45):
                continue
            if dist_km > 0.7 and next_dist > 0.7 and bridge_dist < max(0.3, min(dist_km, next_dist) * 0.55):
                continue

        min_move_km = 0.02 if sensor < 10 else (0.035 if sensor < 40 else 0.06)
        if dist_km < min_move_km and not curr.get("marker_type"):
            continue
        filtered.append(curr)
    return filtered


# --- 차량 관제 (Vehicle Tracking) ---
@app.route("/api/vehicle-tracking", methods=["GET"])
def get_vehicle_tracking():
    """실시간 차량 위치 폴링용 API (Vercel 트래픽 전가 방지)"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        from datetime import timezone, timedelta
        KST = timezone(timedelta(hours=9))
        twenty_four_hours_ago = (datetime.now(KST) - timedelta(hours=24)).isoformat()

        # 최근 24시간 내 운행 중이거나 방금 종료/갱신된 차량 조회
        res = supabase.from_("vehicle_trips") \
                .select("*") \
                .in_("status", ["driving", "paused", "completed"]) \
                .or_(f"started_at.gte.{twenty_four_hours_ago},completed_at.gte.{twenty_four_hours_ago},updated_at.gte.{twenty_four_hours_ago}") \
                .order("started_at", desc=True) \
                .execute()
        
        trips = res.data
        trip_ids = [t["id"] for t in trips]
        
        if not trip_ids:
            return jsonify({"data": [], "trips": []})

        # 각 트립의 마지막 위치 조회 (RPC 대신 Python 루프로 처리 - 나스 리소스 활용)
        loc_res = supabase.from_("vehicle_locations") \
                .select("*") \
                .in_("trip_id", trip_ids) \
                .order("recorded_at", desc=True) \
                .execute()
        
        loc_groups = {}
        for l in loc_res.data:
            loc_groups.setdefault(l["trip_id"], []).append(l)
        loc_map = {}
        for tid, locs in loc_groups.items():
            clean_locs = _vehicle_filter_locations(locs)
            loc_map[tid] = (clean_locs[-1] if clean_locs else locs[0])

        driver_map = {}
        vehicle_numbers = list({t.get("vehicle_number") for t in trips if t.get("vehicle_number")})
        if vehicle_numbers:
            try:
                driver_res = supabase.from_("driver_contacts") \
                    .select("vehicle_number, branch, partner_company, contract_type, cargo_type, map_visibility, general_vehicle_type, general_payload, general_body_type") \
                    .in_("vehicle_number", vehicle_numbers) \
                    .execute()
                for d in (driver_res.data or []):
                    if d.get("vehicle_number"):
                        driver_map[d["vehicle_number"]] = d
            except Exception as meta_err:
                app.logger.warning(f"운전원 메타 조회 실패(무시): {meta_err}")
        
        latest_by_vehicle = {}
        status_rank = {"driving": 0, "paused": 1, "completed": 2}
        for t in trips:
            d = driver_map.get(t.get("vehicle_number")) or {}
            t["cargo_type"] = t.get("cargo_type") or d.get("cargo_type") or "container"
            t["driver_contract_type"] = t.get("driver_contract_type") or d.get("contract_type") or "uncontracted"
            t["map_visibility"] = t.get("map_visibility") or d.get("map_visibility") or "own"
            t["branch"] = t.get("branch") or d.get("branch")
            t["partner_company"] = t.get("partner_company") or d.get("partner_company")
            t["general_vehicle_type"] = t.get("general_vehicle_type") or d.get("general_vehicle_type")
            t["general_payload"] = t.get("general_payload") or d.get("general_payload")
            t["general_body_type"] = t.get("general_body_type") or d.get("general_body_type")
            t["lastLocation"] = loc_map.get(t["id"])
            t["last_location_address"] = t["lastLocation"]["address"] if t["lastLocation"] else None
            vehicle_key = (t.get("vehicle_number") or t.get("vehicle_id") or t.get("id") or "").replace(" ", "").upper()
            last_time = (
                (t.get("lastLocation") or {}).get("recorded_at")
                or (t.get("lastLocation") or {}).get("timestamp")
                or t.get("updated_at")
                or t.get("completed_at")
                or t.get("started_at")
                or ""
            )
            prev = latest_by_vehicle.get(vehicle_key) if vehicle_key else None
            prev_rank = status_rank.get((prev or {}).get("status"), 9)
            curr_rank = status_rank.get(t.get("status"), 9)
            if (
                not vehicle_key
                or not prev
                or curr_rank < prev_rank
                or (curr_rank == prev_rank and last_time > prev.get("_sort_time", ""))
            ):
                t["_sort_time"] = last_time
                latest_by_vehicle[vehicle_key] = t

        merged = sorted(latest_by_vehicle.values(), key=lambda x: x.get("_sort_time", ""), reverse=True)
        merged = sorted(merged, key=lambda x: status_rank.get(x.get("status"), 9))
        for t in merged:
            t.pop("_sort_time", None)

        # 프론트엔드 하위 호환성을 위해 trips와 data 양쪽으로 반환
        return jsonify({"data": merged, "trips": merged})
    except Exception as e:
        app.logger.error(f"관제 데이터 조회 실패: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehicle-tracking/trips/<trip_id>/locations", methods=["GET"])
@app.route("/api/vehicle-tracking/<trip_id>/locations", methods=["GET"])
def get_trip_locations(trip_id):
    """특정 트립의 전체 경로 조회 (나스에서 처리)"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        res = supabase.from_("vehicle_locations") \
                .select("*") \
                .eq("trip_id", trip_id) \
                .order("recorded_at", asc=True) \
                .execute()
        return jsonify({"locations": res.data})
    except Exception as e:
        app.logger.error(f"경로 데이터 조회 실패: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehicle-tracking/trips/<trip_id>", methods=["GET"])
def get_trip_detail(trip_id):
    """트립 상세 정보 조회"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        res = supabase.from_("vehicle_trips").select("*").eq("id", trip_id).single().execute()
        return jsonify(res.data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehicle-tracking/trips/<trip_id>/logs", methods=["GET"])
def get_trip_logs(trip_id):
    """트립 운행 로그 조회"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        res = supabase.from_("vehicle_trip_logs").select("*").eq("trip_id", trip_id).order("created_at", desc=True).execute()
        return jsonify({"logs": res.data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehicle-tracking/trips/<trip_id>", methods=["DELETE"])
def delete_trip(trip_id):
    """트립 삭제"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        # 1. 위치 삭제
        supabase.from_("vehicle_locations").delete().eq("trip_id", trip_id).execute()
        # 2. 로그 삭제
        supabase.from_("vehicle_trip_logs").delete().eq("trip_id", trip_id).execute()
        # 3. 트립 삭제
        supabase.from_("vehicle_trips").delete().eq("id", trip_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehicle-tracking/photos/view", methods=["GET"])
def view_photo():
    """사진 보기 릴레이 (Supabase URL 프록시)"""
    key = request.args.get("key")
    if not key: return "Key missing", 400
    try:
        # Supabase 공용 URL 생성 (버킷 이름: vehicle-photos 가정)
        photo_url = f"{SUPABASE_URL}/storage/v1/object/public/vehicle-photos/{key}"
        return Response(urlopen(photo_url).read(), mimetype="image/jpeg")
    except Exception as e:
        return str(e), 500

@app.route("/api/vehicle-tracking/export/excel", methods=["GET"])
def export_excel_all():
    """전체 기록 엑셀 다운로드"""
    if not supabase: return jsonify({"error": "Supabase not configured"}), 500
    try:
        # 필터 정보
        from_dt = request.args.get("from", "")
        to_dt = request.args.get("to", "")
        keyword = request.args.get("keyword", "")
        
        query = supabase.from_("vehicle_trips").select("*")
        if from_dt: query = query.gte("started_at", f"{from_dt} 00:00:00")
        if to_dt: query = query.lte("started_at", f"{to_dt} 23:59:59")
        
        res = query.order("started_at", desc=True).execute()
        df = pd.DataFrame(res.data)
        
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        
        filename = f"vehicle_records_{datetime.now().strftime('%y%m%d')}.xlsx"
        return Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except Exception as e:
        return str(e), 500

DEFAULT_ASAN_SHIPPING_PATH = "/아산지점/2026_자체보관리스트.xlsx"
SHIPPING_ARCHIVE_RETENTION_DAYS = 365
SHIPPING_LOOKUP_RETENTION_DAYS = 180
shipping_cache = {}
shipping_sync_lock = threading.Lock()
shipping_container_auto_lookup_lock = threading.Lock()
shipping_container_lookup_jobs = {}
shipping_container_lookup_jobs_lock = threading.Lock()
shipping_db_available = True
shipping_history_cleanup_last_date = None
shipping_container_auto_lookup_last_date = None

def resolve_asan_shipping_file(rel_path=None):
    rel_path = (rel_path or DEFAULT_ASAN_SHIPPING_PATH).replace("\\", "/").strip()
    if not rel_path.startswith("/"):
        rel_path = f"/{rel_path}"

    file_path = Path("/app/data") / rel_path.lstrip("/")
    if file_path.exists():
        return file_path, rel_path

    fallback_root = Path("C:/Els") if os.name == "nt" else Path("/Volumes/Els")
    file_path = fallback_root / rel_path.lstrip("/")
    if file_path.exists():
        return file_path, rel_path

    return file_path, rel_path

def parse_asan_shipping_excel(file_path):
    app.logger.info(f"선적관리 엑셀 파싱 시작: {file_path}")
    import shutil, tempfile as _tf
    temp_path = _tf.mktemp(suffix=".xlsx")
    try:
        shutil.copy2(str(file_path), temp_path)
    except Exception as cp_err:
        app.logger.error(f"선적관리 임시복사 실패: {cp_err}")
        raise RuntimeError(f"파일 복사 실패: {cp_err}")

    try:
        df = pd.read_excel(temp_path, sheet_name=0, header=2)
    finally:
        try: os.remove(temp_path)
        except: pass

    raw_cols = [str(c).replace('\n', ' ').strip() for c in df.columns]
    seen = {}
    clean_cols = []
    for c in raw_cols:
        if c.startswith('Unnamed') or c == '':
            c = f'col_{len(clean_cols)+1}'
        if c in seen:
            seen[c] += 1
            c = f'{c}_{seen[c]}'
        else:
            seen[c] = 0
        clean_cols.append(c)
    df.columns = clean_cols

    df = df.fillna("")
    df = df.astype(str).replace(['nan', 'None', '#N/A', 'NaT'], '')

    container_cols = [c for c in df.columns if 'CONTAINER' in c.upper()]
    if container_cols:
        c_col = container_cols[0]
        df = df[df[c_col].str.strip() != ""]

    headers = df.columns.tolist()
    data_rows = df.values.tolist()

    data_rows = [
        [("" if (isinstance(v, float) and math.isnan(v)) or v is None else v) for v in row]
        for row in data_rows
    ]

    return {"headers": headers, "data": data_rows}

def _shipping_value(headers, row, keywords):
    for i, header in enumerate(headers):
        normalized = str(header or "").replace(" ", "").upper()
        if any(keyword.replace(" ", "").upper() in normalized for keyword in keywords):
            return str(row[i] if i < len(row) else "").strip()
    return ""

def _shipping_search_terms(search):
    return [term.strip() for term in str(search or "").split(",") if term.strip()]

def _shipping_search_filter_value(term):
    return str(term).replace(",", " ").replace("\\", " ")

def _shipping_month_keys(months):
    if isinstance(months, (list, tuple, set)):
        raw_months = months
    else:
        raw_months = str(months or "").split(",")
    seen = set()
    keys = []
    for month in raw_months:
        m = re.fullmatch(r"(20\d{2})-(0[1-9]|1[0-2])", str(month or "").strip())
        if not m:
            continue
        key = f"{m.group(1)}-{m.group(2)}"
        if key in seen:
            continue
        seen.add(key)
        keys.append(key)
    return keys

def _shipping_next_month_key(month_key):
    year, month = [int(part) for part in month_key.split("-")]
    if month == 12:
        return f"{year + 1}-01"
    return f"{year}-{month + 1:02d}"

def _shipping_month_ranges(months):
    ranges = [
        {"key": key, "start": f"{key}-01", "end": f"{_shipping_next_month_key(key)}-01", "keys": [key]}
        for key in sorted(_shipping_month_keys(months))
    ]
    merged = []
    for item in ranges:
        if merged and merged[-1]["end"] == item["start"]:
            merged[-1]["end"] = item["end"]
            merged[-1]["keys"].extend(item["keys"])
        else:
            merged.append(item)
    return merged

def _shipping_work_date_index(headers):
    for i, header in enumerate(headers or []):
        text = re.sub(r"\s+", "", str(header or ""))
        if text in ("작업일", "작업일자"):
            return i
    for i, header in enumerate(headers or []):
        text = re.sub(r"\s+", "", str(header or "")).lower()
        if "작업" in text and ("일" in text or "일자" in text):
            return i
    for label in ("반입일", "반입일자", "픽업일", "픽업일자"):
        for i, header in enumerate(headers or []):
            if re.sub(r"\s+", "", str(header or "")) == label:
                return i
    return -1

def _shipping_resolve_date_col(headers, date_col):
    requested = str(date_col or "").strip()
    if requested and requested in (headers or []):
        return requested
    idx = _shipping_work_date_index(headers)
    return headers[idx] if idx >= 0 else ""

def _shipping_apply_month_filter(q, headers, date_col, months):
    resolved_col = _shipping_resolve_date_col(headers, date_col)
    ranges = _shipping_month_ranges(months)
    if not resolved_col or not ranges:
        return q, "", []

    col_expr = f"row_data->>{resolved_col}"
    if len(ranges) == 1:
        item = ranges[0]
        return q.gte(col_expr, item["start"]).lt(col_expr, item["end"]), resolved_col, item["keys"]

    filters = ",".join(
        f"and({col_expr}.gte.{item['start']},{col_expr}.lt.{item['end']})"
        for item in ranges
    )
    keys = [key for item in ranges for key in item["keys"]]
    return q.or_(filters), resolved_col, keys

def _shipping_rows_query(normalized_path, headers, search_terms, date_col, months, count=None):
    if count:
        q = supabase.from_("branch_shipping_rows").select("row_values,row_index", count=count)
    else:
        q = supabase.from_("branch_shipping_rows").select("row_values,row_index")
    q = q.eq("branch_id", "asan").eq("file_path", normalized_path)
    if len(search_terms) == 1:
        q = q.ilike("search_text", f"%{_shipping_search_filter_value(search_terms[0])}%")
    elif search_terms:
        filters = ",".join(f"search_text.ilike.%{_shipping_search_filter_value(term)}%" for term in search_terms)
        q = q.or_(filters)
    return _shipping_apply_month_filter(q, headers, date_col, months)

def _shipping_fetch_rows_in_chunks(query_factory, start, end, chunk_size=1000):
    rows = []
    total = None
    cursor = max(0, int(start or 0))
    final_end = max(cursor, int(end or cursor))

    while cursor <= final_end:
        chunk_end = min(final_end, cursor + chunk_size - 1)
        q, _, _ = query_factory(count="exact" if total is None else None)
        rows_res = q.order("row_index", desc=False).range(cursor, chunk_end).execute()
        chunk_rows = rows_res.data or []
        rows.extend(chunk_rows)

        if total is None and rows_res.count is not None:
            total = rows_res.count
            final_end = min(final_end, max(0, total - 1))
        if not chunk_rows or (total is None and len(chunk_rows) < (chunk_end - cursor + 1)):
            break
        cursor = chunk_end + 1

    return rows, total

def _shipping_normalize_container_no(value):
    return re.sub(r"\s+", "", str(value or "")).upper()

def _shipping_is_container_no_shape(value):
    return bool(re.fullmatch(r"[A-Z]{4}\d{7}", _shipping_normalize_container_no(value)))

def _shipping_is_actual_history_row(row):
    if not isinstance(row, (list, tuple)) or len(row) < 4:
        return False
    if not re.fullmatch(r"\d+", str(row[1] or "").strip()):
        return False
    text = "|".join(str(cell or "") for cell in row[2:14])
    return any(token in text for token in ("수입", "수출", "반입", "반출", "양하", "적하"))

def _shipping_lookup_main_status(record):
    main_row = record.get("main_row") if isinstance(record, dict) else None
    if isinstance(main_row, list) and len(main_row) > 3:
        return str(main_row[3] or "").strip()
    return str((record or {}).get("main_status") or "").strip()

def _shipping_latest_lookup_statuses(normalized_path, containers):
    statuses = {}
    if not supabase or not containers:
        return statuses
    ordered = [_shipping_normalize_container_no(cn) for cn in containers if _shipping_is_container_no_shape(cn)]
    for i in range(0, len(ordered), 500):
        chunk = ordered[i:i + 500]
        try:
            res = supabase.from_("branch_shipping_container_lookups") \
                .select("container_no,main_status,main_row,looked_up_at") \
                .eq("branch_id", "asan") \
                .eq("file_path", normalized_path) \
                .in_("container_no", chunk) \
                .order("looked_up_at", desc=True) \
                .execute()
        except Exception as exc:
            app.logger.warning(f"[컨테이너자동조회] 기존 이력 조회 실패: {exc}")
            return statuses
        for item in res.data or []:
            cn = _shipping_normalize_container_no(item.get("container_no"))
            if cn and cn not in statuses:
                statuses[cn] = _shipping_lookup_main_status(item)
    return statuses

def _shipping_container_auto_lookup_enabled():
    try:
        settings = get_asan_dispatch_settings(force=True) or {}
        if "shipping_container_auto_lookup_enabled" not in settings:
            app.logger.warning("[컨테이너자동조회] DB 설정 컬럼 미적용 상태라 자동조회 실행을 보류")
            return False
        return settings.get("shipping_container_auto_lookup_enabled", True) is not False
    except Exception as exc:
        app.logger.warning(f"[컨테이너자동조회] 설정 조회 실패로 자동조회 실행 보류: {exc}")
        return False

def set_asan_shipping_container_auto_lookup_enabled(enabled, reason=""):
    if not supabase:
        return False
    try:
        settings = get_asan_dispatch_settings(force=True) or {}
        payload = {
            "shipping_container_auto_lookup_enabled": bool(enabled),
            "updated_at": datetime.now(KST).isoformat(),
        }
        supabase.from_("branch_dispatch_settings").update(payload).eq("branch_id", "asan").execute()
        dispatch_settings_cache["data"] = {**settings, **payload}
        dispatch_settings_cache["loaded_at"] = time.time()
        suffix = f" ({reason})" if reason else ""
        app.logger.warning(f"[컨테이너자동조회] 사용 여부 {bool(enabled)} 저장{suffix}")
        return True
    except Exception as exc:
        app.logger.warning(f"[컨테이너자동조회] 사용 여부 저장 실패: {exc}")
        return False

def _shipping_auto_lookup_targets(rel_path=None, max_targets=ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MAX_TARGETS):
    normalized_path = (rel_path or DEFAULT_ASAN_SHIPPING_PATH).replace("\\", "/").strip()
    if not normalized_path.startswith("/"):
        normalized_path = f"/{normalized_path}"

    meta_res = supabase.from_("branch_shipping_files").select("*").eq("branch_id", "asan").eq("file_path", normalized_path).execute()
    meta = meta_res.data[0] if meta_res.data else None
    if not meta:
        return normalized_path, [], 0, 0

    headers = meta.get("headers") or []
    container_idx = -1
    for i, header in enumerate(headers):
        text = str(header or "").upper()
        if "CONTAINER" in text or "컨테이너" in text:
            container_idx = i
            break

    rows = []
    cursor = 0
    chunk_size = 1000
    while len(rows) < max_targets:
        end = min(cursor + chunk_size - 1, max_targets - 1)
        res = supabase.from_("branch_shipping_rows") \
            .select("row_values,row_index,container_no") \
            .eq("branch_id", "asan") \
            .eq("file_path", normalized_path) \
            .order("row_index", desc=False) \
            .range(cursor, end) \
            .execute()
        chunk = res.data or []
        rows.extend(chunk)
        if len(chunk) < (end - cursor + 1):
            break
        cursor = end + 1

    seen = set()
    containers = []
    for item in rows:
        row = item.get("row_values") or []
        cn = _shipping_normalize_container_no(item.get("container_no") or (row[container_idx] if 0 <= container_idx < len(row) else ""))
        if not _shipping_is_container_no_shape(cn) or cn in seen:
            continue
        seen.add(cn)
        containers.append(cn)

    statuses = _shipping_latest_lookup_statuses(normalized_path, containers)
    targets = [cn for cn in containers if statuses.get(cn) != "적하"]
    skipped = len(containers) - len(targets)
    return normalized_path, targets, len(containers), skipped

def _shipping_group_actual_lookup_rows(rows):
    grouped = {}
    for row in rows or []:
        if not _shipping_is_actual_history_row(row):
            continue
        cn = _shipping_normalize_container_no(row[0] if row else "")
        if not cn:
            continue
        grouped.setdefault(cn, []).append(list(row))
    for cn in list(grouped):
        grouped[cn].sort(key=lambda item: int(item[1]) if str(item[1]).isdigit() else 999)
    return grouped

def _save_asan_shipping_container_lookup_rows(normalized_path, rows, lookup_source="asan_shipping_auto"):
    grouped = _shipping_group_actual_lookup_rows(rows)
    if not grouped:
        return 0

    containers = list(grouped.keys())
    supabase.from_("branch_shipping_container_lookups") \
        .delete() \
        .eq("branch_id", "asan") \
        .eq("file_path", normalized_path) \
        .in_("container_no", containers) \
        .execute()

    looked_up_at = datetime.now(KST).isoformat()
    run_id = str(uuid.uuid4())
    payload = []
    for cn, container_rows in grouped.items():
        main_row = next((row for row in container_rows if str(row[1]).strip() == "1"), container_rows[0])
        payload.append({
            "run_id": run_id,
            "branch_id": "asan",
            "file_path": normalized_path,
            "container_no": cn,
            "result_rows": container_rows,
            "main_row": main_row,
            "main_status": str(main_row[2] if len(main_row) > 2 else ""),
            "terminal": str(main_row[4] if len(main_row) > 4 else ""),
            "move_time": str(main_row[5] if len(main_row) > 5 else ""),
            "vehicle_no": str(main_row[13] if len(main_row) > 13 else ""),
            "lookup_source": lookup_source,
            "looked_up_at": looked_up_at,
            "updated_at": looked_up_at,
        })
    supabase.from_("branch_shipping_container_lookups").insert(payload).execute()
    return len(payload)

def _shipping_lookup_rows_failed(rows):
    if not rows:
        return True
    first = rows[0] if isinstance(rows[0], (list, tuple)) else []
    return str(first[1] if len(first) > 1 else "").strip().upper() == "ERROR"

def _shipping_normalize_path(rel_path=None):
    normalized_path = (rel_path or DEFAULT_ASAN_SHIPPING_PATH).replace("\\", "/").strip()
    if not normalized_path.startswith("/"):
        normalized_path = f"/{normalized_path}"
    return normalized_path

def _shipping_lookup_failure_reason(rows):
    if not rows:
        return "결과 없음"
    first = rows[0] if isinstance(rows[0], (list, tuple)) else []
    return str(first[2] if len(first) > 2 else "조회 실패").strip() or "조회 실패"

def _shipping_lookup_job_snapshot(job):
    error_reasons = job.get("error_reasons") or {}
    reasons = [
        {"reason": reason, "count": count}
        for reason, count in sorted(error_reasons.items(), key=lambda item: item[1], reverse=True)
    ]
    total = int(job.get("total") or 0)
    completed = len(job.get("completed_set") or set())
    failed = len((job.get("failed_set") or set()) - (job.get("completed_set") or set()))
    return {
        "id": job.get("id"),
        "path": job.get("path"),
        "state": job.get("state"),
        "status": job.get("status"),
        "total": total,
        "completed": completed,
        "failed": failed,
        "remaining": max(0, total - completed - failed),
        "saved": int(job.get("saved") or 0),
        "startedAt": job.get("started_at"),
        "updatedAt": job.get("updated_at"),
        "errorSummary": {
            "total": sum(error_reasons.values()),
            "reasons": reasons,
            "message": ", ".join(f"{item['reason']} {item['count']}건" for item in reasons[:2]),
        } if reasons else None,
    }

def _update_shipping_lookup_job(job_id, **patch):
    with shipping_container_lookup_jobs_lock:
        job = shipping_container_lookup_jobs.get(job_id)
        if not job:
            return None
        job.update(patch)
        job["updated_at"] = datetime.now(KST).isoformat()
        return _shipping_lookup_job_snapshot(job)

def _apply_shipping_lookup_job_rows(job_id, normalized_path, rows):
    cn = _shipping_normalize_container_no(rows[0][0] if rows and rows[0] else "")
    failed = _shipping_lookup_rows_failed(rows)
    saved = 0
    reason = ""
    if failed:
        reason = _shipping_lookup_failure_reason(rows)
    else:
        try:
            saved = _save_asan_shipping_container_lookup_rows(
                normalized_path,
                rows,
                lookup_source="asan_shipping_manual_background",
            )
        except Exception as exc:
            failed = True
            reason = f"저장 실패: {exc}"

    with shipping_container_lookup_jobs_lock:
        job = shipping_container_lookup_jobs.get(job_id)
        if not job:
            return None
        if cn:
            if failed:
                if cn not in job["completed_set"]:
                    job["failed_set"].add(cn)
                job["error_reasons"][reason] = job["error_reasons"].get(reason, 0) + 1
            else:
                job["completed_set"].add(cn)
                job["failed_set"].discard(cn)
                job["saved"] = int(job.get("saved") or 0) + saved
        snapshot = _shipping_lookup_job_snapshot(job)
        job["status"] = (
            f"백그라운드 조회 중: 완료 {snapshot['completed']}건 / 실패 {snapshot['failed']}건 / "
            f"대상 {snapshot['total']}건"
        )
        job["updated_at"] = datetime.now(KST).isoformat()
        return snapshot

def _run_shipping_container_lookup_job(job_id):
    response = None
    with shipping_container_lookup_jobs_lock:
        job = shipping_container_lookup_jobs.get(job_id)
        if not job:
            return
        normalized_path = job["path"]
        containers = list(job["containers"])

    try:
        _update_shipping_lookup_job(job_id, state="running", status=f"백그라운드 컨테이너 조회 시작: {len(containers)}건")
        try:
            cap = requests.get(f"{ELS_BOT_API_URL}/api/els/capabilities", timeout=5).json()
            if int(cap.get("total_drivers") or 0) <= 0:
                requests.post(f"{ELS_BOT_API_URL}/api/els/warmup", json={"wait": False}, timeout=10)
        except Exception as exc:
            app.logger.warning(f"[컨테이너백그라운드조회] 봇 준비 확인 실패: {exc}")

        response = requests.post(
            f"{ELS_BOT_API_URL}/api/els/run",
            json={
                "containers": containers,
                "reserveSingle": False,
                "stableBatchMode": len(containers) >= 100,
                "maxBatchWorkers": 1 if len(containers) >= 100 else None,
                "acquireTimeoutSec": 300 if len(containers) >= 100 else 45,
                "submitDelaySec": 2 if len(containers) >= 100 else 0,
            },
            stream=True,
            timeout=(10, max(3600, len(containers) * 30)),
        )
        response.raise_for_status()
        response.encoding = "utf-8"
        for line in response.iter_lines(decode_unicode=True):
            if not line:
                continue
            with shipping_container_lookup_jobs_lock:
                stop_requested = bool((shipping_container_lookup_jobs.get(job_id) or {}).get("stop_requested"))
            if stop_requested:
                _update_shipping_lookup_job(job_id, state="cancelled", status="사용자 요청으로 백그라운드 조회 중단")
                try:
                    requests.post(f"{ELS_BOT_API_URL}/api/els/stop-daemon", timeout=5)
                except Exception:
                    pass
                return
            if line.startswith("LOG:"):
                _update_shipping_lookup_job(job_id, status=line[4:].strip())
                continue
            if line.startswith("RESULT_PARTIAL:"):
                payload = json.loads(line[15:])
                rows = payload.get("result") or []
                _apply_shipping_lookup_job_rows(job_id, normalized_path, rows)
                continue
            if line.startswith("RESULT:"):
                payload = json.loads(line[7:])
                if payload.get("ok") is False:
                    raise RuntimeError(payload.get("error") or "컨테이너 조회 실패")

        snapshot = _update_shipping_lookup_job(job_id)
        final_state = "failed" if snapshot and snapshot["failed"] > 0 else "completed"
        final_status = (
            f"백그라운드 조회 종료: 완료 {snapshot['completed']}건, 실패 {snapshot['failed']}건"
            if snapshot else "백그라운드 조회 종료"
        )
        _update_shipping_lookup_job(job_id, state=final_state, status=final_status)
    except Exception as exc:
        app.logger.error(f"[컨테이너백그라운드조회] 실행 오류: {exc}", exc_info=True)
        _update_shipping_lookup_job(job_id, state="failed", status=f"백그라운드 조회 오류: {exc}")
    finally:
        if response is not None:
            try:
                response.close()
            except Exception:
                pass

@app.route("/api/branches/asan/shipping/container-lookup/jobs", methods=["GET", "POST", "DELETE"])
def asan_shipping_container_lookup_jobs():
    if not supabase or not shipping_db_available:
        return jsonify({"ok": False, "error": "Supabase/선적관리 DB 미사용 상태입니다."}), 503

    if request.method == "GET":
        job_id = request.args.get("id")
        with shipping_container_lookup_jobs_lock:
            job = shipping_container_lookup_jobs.get(job_id) if job_id else None
            if not job:
                return jsonify({"ok": False, "error": "작업을 찾을 수 없습니다."}), 404
            return jsonify({"ok": True, "job": _shipping_lookup_job_snapshot(job)})

    data = request.get_json(silent=True) or {}
    if request.method == "DELETE":
        job_id = data.get("id") or request.args.get("id")
        with shipping_container_lookup_jobs_lock:
            job = shipping_container_lookup_jobs.get(job_id)
            if not job:
                return jsonify({"ok": False, "error": "작업을 찾을 수 없습니다."}), 404
            job["stop_requested"] = True
            job["state"] = "stopping"
            job["status"] = "사용자 중지 요청 처리 중"
            job["updated_at"] = datetime.now(KST).isoformat()
            snapshot = _shipping_lookup_job_snapshot(job)
        try:
            requests.post(f"{ELS_BOT_API_URL}/api/els/stop-daemon", timeout=5)
        except Exception:
            pass
        return jsonify({"ok": True, "job": snapshot})

    normalized_path = _shipping_normalize_path(data.get("path") or data.get("file_path"))
    seen = set()
    containers = []
    for value in data.get("containers") or []:
        cn = _shipping_normalize_container_no(value)
        if not _shipping_is_container_no_shape(cn) or cn in seen:
            continue
        seen.add(cn)
        containers.append(cn)
    if not containers:
        return jsonify({"ok": False, "error": "조회할 컨테이너가 없습니다."}), 400

    with shipping_container_lookup_jobs_lock:
        for job in shipping_container_lookup_jobs.values():
            if job.get("state") in ("running", "stopping"):
                return jsonify({"ok": True, "already_running": True, "job": _shipping_lookup_job_snapshot(job)}), 202
        job_id = str(uuid.uuid4())
        now_iso = datetime.now(KST).isoformat()
        shipping_container_lookup_jobs[job_id] = {
            "id": job_id,
            "path": normalized_path,
            "containers": containers,
            "total": len(containers),
            "state": "running",
            "status": f"백그라운드 컨테이너 조회 준비: {len(containers)}건",
            "saved": 0,
            "completed_set": set(),
            "failed_set": set(),
            "error_reasons": {},
            "stop_requested": False,
            "started_at": now_iso,
            "updated_at": now_iso,
        }
        snapshot = _shipping_lookup_job_snapshot(shipping_container_lookup_jobs[job_id])

    threading.Thread(target=_run_shipping_container_lookup_job, args=(job_id,), daemon=True).start()
    return jsonify({"ok": True, "job": snapshot}), 202

def run_asan_shipping_container_auto_lookup():
    if not shipping_container_auto_lookup_lock.acquire(blocking=False):
        app.logger.info("[컨테이너자동조회] 이미 실행 중이라 건너뜀")
        return
    try:
        if not _shipping_container_auto_lookup_enabled():
            app.logger.info("[컨테이너자동조회] 설정 OFF 상태라 실행하지 않음")
            return
        if not supabase or not shipping_db_available:
            app.logger.warning("[컨테이너자동조회] Supabase/선적관리 DB 미사용 상태라 실행 불가")
            return

        normalized_path, containers, total_containers, skipped_loaded = _shipping_auto_lookup_targets()
        app.logger.info(
            f"[컨테이너자동조회] 시작: 전체 컨테이너 {total_containers}건, 적하 제외 {skipped_loaded}건, 조회대상 {len(containers)}건"
        )
        if not containers:
            return

        try:
            cap = requests.get(f"{ELS_BOT_API_URL}/api/els/capabilities", timeout=5).json()
            if cap.get("progress", {}).get("is_running"):
                app.logger.warning("[컨테이너자동조회] 기존 컨테이너 조회가 진행 중이라 오늘 자동조회는 건너뜀")
                return
        except Exception as exc:
            app.logger.warning(f"[컨테이너자동조회] 봇 상태 확인 실패: {exc}")

        failed_count = 0
        completed_count = 0
        saved_count = 0
        response = None
        try:
            response = requests.post(
                f"{ELS_BOT_API_URL}/api/els/run",
                json={
                    "containers": containers,
                    "reserveSingle": False,
                    "stableBatchMode": True,
                    "maxBatchWorkers": 1,
                    "acquireTimeoutSec": ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_ACQUIRE_TIMEOUT_SECONDS,
                    "submitDelaySec": ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_SUBMIT_DELAY_SECONDS,
                },
                stream=True,
                timeout=(10, ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_TIMEOUT_SECONDS),
            )
            response.raise_for_status()
            response.encoding = "utf-8"
            for line in response.iter_lines(decode_unicode=True):
                if not line:
                    continue
                if line.startswith("LOG:"):
                    app.logger.info(f"[컨테이너자동조회] {line[4:].strip()}")
                    continue
                if line.startswith("RESULT_PARTIAL:"):
                    payload = json.loads(line[15:])
                    rows = payload.get("result") or []
                    if _shipping_lookup_rows_failed(rows):
                        failed_count += 1
                    else:
                        completed_count += 1
                        saved_count += _save_asan_shipping_container_lookup_rows(normalized_path, rows)
                    if failed_count >= ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_FAIL_LIMIT:
                        reason = f"조회 실패 {failed_count}회 도달"
                        set_asan_shipping_container_auto_lookup_enabled(False, reason=reason)
                        app.logger.error(f"[컨테이너자동조회] {reason}로 자동조회 중지")
                        try:
                            if response is not None:
                                response.close()
                        finally:
                            try:
                                requests.post(f"{ELS_BOT_API_URL}/api/els/stop-daemon", timeout=5)
                            except Exception:
                                pass
                        return
                elif line.startswith("RESULT:"):
                    payload = json.loads(line[7:])
                    if payload.get("ok") is False:
                        raise RuntimeError(payload.get("error") or "컨테이너 자동조회 실패")
            app.logger.info(
                f"[컨테이너자동조회] 완료: 조회완료 {completed_count}건, 조회실패 {failed_count}건, 저장 {saved_count}건"
            )
        except Exception as exc:
            set_asan_shipping_container_auto_lookup_enabled(False, reason=f"실행 오류: {exc}")
            app.logger.error(f"[컨테이너자동조회] 실행 오류로 중지: {exc}", exc_info=True)
        finally:
            if response is not None:
                try:
                    response.close()
                except Exception:
                    pass
    finally:
        shipping_container_auto_lookup_lock.release()

def maybe_run_asan_shipping_container_auto_lookup(now=None):
    global shipping_container_auto_lookup_last_date
    now = now or datetime.now(KST)
    today_key = now.date().isoformat()
    minute_end = min(60, ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MINUTE + 2)
    in_window = (
        now.hour == ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_HOUR
        and ASAN_SHIPPING_CONTAINER_AUTO_LOOKUP_MINUTE <= now.minute < minute_end
    )
    if not in_window or shipping_container_auto_lookup_last_date == today_key:
        return
    shipping_container_auto_lookup_last_date = today_key
    threading.Thread(target=run_asan_shipping_container_auto_lookup, daemon=True).start()

def _shipping_sort_value(value):
    s = str(value if value is not None else "").strip()
    if not s:
        return None

    compact = s.replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", compact):
        return (0, float(compact))

    date_match = re.fullmatch(r"(\d{4})[-/.]?(\d{2})[-/.]?(\d{2})(?:\.0)?", s)
    if date_match:
        return (1, tuple(int(part) for part in date_match.groups()))

    time_match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::\d{2})?(?:\.\d+)?", s)
    if time_match:
        return (2, int(time_match.group(1)), int(time_match.group(2)))

    return (3, s.casefold())

def _shipping_row_hash(row):
    return hashlib.sha256(json.dumps(row or [], ensure_ascii=False, default=str).encode("utf-8")).hexdigest()

def _archive_removed_asan_shipping_rows(normalized_path, new_payload, removed_file_modified_at):
    try:
        existing_res = supabase.from_("branch_shipping_rows").select(
            "row_index,row_values,row_data,container_no,vessel_name,search_text,file_modified_at"
        ).eq("branch_id", "asan").eq("file_path", normalized_path).execute()
        existing_rows = existing_res.data or []
        if not existing_rows:
            return 0

        new_hashes = {_shipping_row_hash(item.get("row_values")) for item in new_payload}
        archived_at = datetime.now(KST).isoformat()
        archive_payload = []
        for row in existing_rows:
            row_hash = _shipping_row_hash(row.get("row_values") or [])
            if row_hash in new_hashes:
                continue
            archive_payload.append({
                "branch_id": "asan",
                "file_path": normalized_path,
                "row_index": row.get("row_index"),
                "row_values": row.get("row_values") or [],
                "row_data": row.get("row_data") or {},
                "container_no": row.get("container_no") or "",
                "vessel_name": row.get("vessel_name") or "",
                "search_text": row.get("search_text") or "",
                "source_row_hash": row_hash,
                "original_file_modified_at": row.get("file_modified_at"),
                "removed_file_modified_at": removed_file_modified_at,
                "archive_reason": "deleted_from_excel",
                "archived_at": archived_at,
            })

        for i in range(0, len(archive_payload), 500):
            supabase.from_("branch_shipping_row_archive").insert(archive_payload[i:i + 500]).execute()
        return len(archive_payload)
    except Exception as e:
        app.logger.warning(f"[선적관리DB] 삭제 이력 archive 건너뜀: {e}")
        return 0

def cleanup_asan_shipping_history_retention(
    now=None,
    archive_retention_days=SHIPPING_ARCHIVE_RETENTION_DAYS,
    lookup_retention_days=SHIPPING_LOOKUP_RETENTION_DAYS
):
    if not supabase or not shipping_db_available:
        return None

    now = now or datetime.now(KST)
    archive_cutoff = (now - timedelta(days=archive_retention_days)).isoformat()
    lookup_cutoff = (now - timedelta(days=lookup_retention_days)).isoformat()
    try:
        archive_res = supabase.from_("branch_shipping_row_archive").delete().lt("archived_at", archive_cutoff).execute()
        lookup_res = supabase.from_("branch_shipping_container_lookups").delete().lt("looked_up_at", lookup_cutoff).execute()
        archive_deleted = len(archive_res.data or [])
        lookup_deleted = len(lookup_res.data or [])
        app.logger.info(
            f"[선적관리DB] 이력 보존기간 정리 완료: archive_cutoff={archive_cutoff}, lookup_cutoff={lookup_cutoff}, "
            f"archive={archive_deleted}, lookup={lookup_deleted}"
        )
        return {
            "archive_cutoff": archive_cutoff,
            "lookup_cutoff": lookup_cutoff,
            "archive_deleted": archive_deleted,
            "lookup_deleted": lookup_deleted,
        }
    except Exception as e:
        app.logger.warning(f"[선적관리DB] 이력 보존기간 정리 실패: {e}")
        return None

def maybe_cleanup_asan_shipping_history(now=None):
    global shipping_history_cleanup_last_date
    now = now or datetime.now(KST)
    today_key = now.date().isoformat()
    if shipping_history_cleanup_last_date == today_key:
        return None
    if now.hour < 3:
        return None

    shipping_history_cleanup_last_date = today_key
    return cleanup_asan_shipping_history_retention(now=now)

def sync_asan_shipping_python(force=False, rel_path=None):
    global shipping_db_available
    if not supabase or not shipping_db_available:
        return None

    file_path, normalized_path = resolve_asan_shipping_file(rel_path)
    if not file_path.exists():
        app.logger.warning(f"[선적관리DB] 파일을 찾을 수 없음: {file_path}")
        return None

    file_stat = file_path.stat()
    mtime_ts = file_stat.st_mtime
    file_modified_at = datetime.fromtimestamp(mtime_ts, tz=KST).isoformat()
    file_signature = (mtime_ts, file_stat.st_size)
    parsed = None
    rows = None
    payload = None
    lock_acquired = shipping_sync_lock.acquire(blocking=False)
    if not lock_acquired:
        app.logger.info(f"[선적관리DB] 동기화 이미 진행 중: {normalized_path}")
        try:
            meta_res = supabase.from_("branch_shipping_files").select("*").eq("branch_id", "asan").eq("file_path", normalized_path).execute()
            if meta_res.data:
                current_meta = meta_res.data[0]
                current_meta["sync_running"] = True
                return current_meta
        except Exception:
            pass
        return {"file_modified_at": file_modified_at, "sync_running": True}

    try:
        meta_res = supabase.from_("branch_shipping_files").select("file_modified_at").eq("branch_id", "asan").eq("file_path", normalized_path).execute()
        current_meta = meta_res.data[0] if meta_res.data else None
        if not force and current_meta and current_meta.get("file_modified_at"):
            try:
                db_mtime = datetime.fromisoformat(current_meta["file_modified_at"].replace("Z", "+00:00")).timestamp()
                if abs(db_mtime - mtime_ts) < 1:
                    shipping_sync_gate.mark_synced(normalized_path, file_signature)
                    return current_meta
            except Exception:
                pass

        decision = shipping_sync_gate.check(normalized_path, file_signature, force=force)
        if not decision.ready:
            if decision.reason in ("pending", "settling"):
                app.logger.info(f"[선적관리DB] 파일 저장 안정화 대기 중: {normalized_path} ({decision.reason})")
            return current_meta

        parsed = parse_asan_shipping_excel(file_path)
        headers = parsed["headers"]
        rows = parsed["data"]

        payload = []
        for row_index, row in enumerate(rows):
            row_data = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
            search_text = " ".join(str(v) for v in row if v)
            payload.append({
                "branch_id": "asan",
                "file_path": normalized_path,
                "row_index": row_index,
                "row_values": row,
                "row_data": row_data,
                "container_no": _shipping_value(headers, row, ["CONTAINER"]),
                "vessel_name": _shipping_value(headers, row, ["선적확정모선", "모선", "VESSEL"]),
                "search_text": search_text[:8000],
                "file_modified_at": file_modified_at,
                "updated_at": datetime.now(KST).isoformat()
            })

        archived_count = _archive_removed_asan_shipping_rows(normalized_path, payload, file_modified_at)
        supabase.from_("branch_shipping_rows").delete().eq("branch_id", "asan").eq("file_path", normalized_path).execute()
        for i in range(0, len(payload), 500):
            supabase.from_("branch_shipping_rows").insert(payload[i:i + 500]).execute()

        supabase.from_("branch_shipping_files").upsert({
            "branch_id": "asan",
            "file_path": normalized_path,
            "headers": headers,
            "row_count": len(rows),
            "file_modified_at": file_modified_at,
            "synced_at": datetime.now(KST).isoformat()
        }, on_conflict="branch_id,file_path").execute()

        shipping_cache.pop(normalized_path, None)
        shipping_sync_gate.mark_synced(normalized_path, file_signature)
        app.logger.info(f"[선적관리DB] 동기화 완료: {normalized_path} ({len(rows)}행, 삭제 archive {archived_count}행)")
        return {"file_modified_at": file_modified_at, "archived_count": archived_count}
    except Exception as e:
        if "branch_shipping_" in str(e) or "relation" in str(e).lower():
            shipping_db_available = False
            app.logger.warning("[선적관리DB] Supabase 테이블이 없어 DB 동기화를 비활성화합니다. migration 적용 후 컨테이너를 재시작하세요.")
        else:
            app.logger.error(f"[선적관리DB] 동기화 실패: {e}", exc_info=True)
        return None
    finally:
        parsed = None
        rows = None
        payload = None
        if lock_acquired:
            shipping_sync_lock.release()
        gc.collect()

def query_asan_shipping_db(rel_path, page=1, page_size=5000, search="", sort_key="", sort_dir="asc", date_col="", months=""):
    if not supabase or not shipping_db_available:
        return None

    normalized_path = (rel_path or DEFAULT_ASAN_SHIPPING_PATH).replace("\\", "/").strip()
    if not normalized_path.startswith("/"):
        normalized_path = f"/{normalized_path}"

    meta_res = supabase.from_("branch_shipping_files").select("*").eq("branch_id", "asan").eq("file_path", normalized_path).execute()
    meta = meta_res.data[0] if meta_res.data else None
    if not meta:
        return None

    page = max(1, int(page or 1))
    page_size = max(1, min(10000, int(page_size or 5000)))
    start = (page - 1) * page_size
    end = start + page_size - 1

    headers = meta.get("headers") or []
    sort_key = str(sort_key or "").strip()
    sort_desc = str(sort_dir or "asc").lower() == "desc"
    sort_idx = headers.index(sort_key) if sort_key in headers else -1

    search_terms = _shipping_search_terms(search)
    def query_factory(count=None):
        return _shipping_rows_query(normalized_path, headers, search_terms, date_col, months, count=count)

    _, resolved_date_col, filtered_months = query_factory()

    if sort_idx >= 0:
        fetched_rows, total_count = _shipping_fetch_rows_in_chunks(query_factory, 0, 9999)
        sortable = []
        blanks = []
        for item in fetched_rows:
            row = item.get("row_values") or []
            sort_value = _shipping_sort_value(row[sort_idx] if sort_idx < len(row) else "")
            if sort_value is None:
                blanks.append(item)
            else:
                sortable.append((sort_value, item))
        sortable.sort(key=lambda pair: pair[0], reverse=sort_desc)
        ordered_rows = ([item for _, item in sortable] + blanks) if sort_desc else (blanks + [item for _, item in sortable])
        page_rows = ordered_rows[start:end + 1]
    else:
        page_rows, total_count = _shipping_fetch_rows_in_chunks(query_factory, start, end)

    return {
        "headers": headers,
        "data": [r.get("row_values") for r in page_rows],
        "file_modified_at": meta.get("file_modified_at"),
        "synced_at": meta.get("synced_at"),
        "total": total_count if total_count is not None else meta.get("row_count", 0),
        "page": page,
        "page_size": page_size,
        "sort_key": sort_key if sort_idx >= 0 else "",
        "sort_dir": "desc" if sort_desc else "asc",
        "date_col": resolved_date_col,
        "months": filtered_months,
        "source": "supabase"
    }

def asan_shipping_sync_scheduler():
    app.logger.info(
        f"[스케줄러] 아산 선적관리 DB 동기화 스케줄러 시작 "
        f"(poll={ASAN_SHIPPING_SYNC_POLL_SECONDS}s, quiet={ASAN_SHIPPING_SYNC_QUIET_SECONDS}s)"
    )
    while True:
        try:
            now = datetime.now(KST)
            maybe_cleanup_asan_shipping_history(now)
            maybe_run_asan_shipping_container_auto_lookup(now)
            if 6 <= now.hour <= 23:
                sync_asan_shipping_python()
            time.sleep(ASAN_SHIPPING_SYNC_POLL_SECONDS)
        except Exception as e:
            app.logger.error(f"[선적관리DB 스케줄러] 에러: {e}")
            time.sleep(ASAN_SHIPPING_SYNC_POLL_SECONDS)

threading.Thread(target=asan_shipping_sync_scheduler, daemon=True).start()

@app.route("/api/branches/asan/shipping", methods=["GET", "POST"])
def get_asan_shipping():
    """아산지점 선적관리 조회/동기화.

    GET은 빠른 조회를 위해 Supabase DB를 먼저 읽고, 요청 중 엑셀 파싱/동기화를 하지 않는다.
    POST는 사용자가 명시적으로 누른 NAS 동기화 버튼과 운영 스케줄러용 강제 동기화 경로다.
    """
    try:
        rel_path = request.args.get("path") or DEFAULT_ASAN_SHIPPING_PATH
        body = request.get_json(silent=True) or {}
        if request.method == "POST":
            rel_path = body.get("path") or rel_path
            force = body.get("force", True)
            if isinstance(force, str):
                force = force.lower() in ("1", "true", "yes", "y")

            if not supabase or not shipping_db_available:
                return jsonify({"ok": False, "error": "선적관리 Supabase 동기화가 비활성화되어 있습니다."}), 503

            sync_result = sync_asan_shipping_python(force=bool(force), rel_path=rel_path)
            if not sync_result:
                return jsonify({"ok": False, "error": "선적관리 NAS 동기화에 실패했습니다."}), 500

            page = body.get("page") or request.args.get("page", 1)
            page_size = body.get("page_size") or request.args.get("page_size", 5000)
            search = (body.get("search") or request.args.get("search") or "").strip()
            sort_key = (body.get("sort_key") or request.args.get("sort_key") or "").strip()
            sort_dir = body.get("sort_dir") or request.args.get("sort_dir") or "asc"
            date_col = (body.get("date_col") or request.args.get("date_col") or "").strip()
            months = body.get("months") or request.args.get("months") or ""
            db_data = query_asan_shipping_db(
                rel_path,
                page=page,
                page_size=page_size,
                search=search,
                sort_key=sort_key,
                sort_dir=sort_dir,
                date_col=date_col,
                months=months
            )
            return jsonify({"ok": True, "data": db_data or sync_result})

        source = request.args.get("source", "auto")
        page = request.args.get("page", 1)
        page_size = request.args.get("page_size", 5000)
        search = (request.args.get("search") or "").strip()
        sort_key = (request.args.get("sort_key") or "").strip()
        sort_dir = request.args.get("sort_dir") or "asc"
        date_col = (request.args.get("date_col") or "").strip()
        months = request.args.get("months") or ""

        if source != "excel" and supabase and shipping_db_available:
            db_data = query_asan_shipping_db(
                rel_path,
                page=page,
                page_size=page_size,
                search=search,
                sort_key=sort_key,
                sort_dir=sort_dir,
                date_col=date_col,
                months=months
            )
            if db_data:
                return jsonify({"data": db_data})

        file_path, normalized_path = resolve_asan_shipping_file(rel_path)
        if not file_path.exists():
            return jsonify({"error": "선적관리 엑셀 파일을 찾을 수 없습니다."}), 404

        mtime = file_path.stat().st_mtime
        cached = shipping_cache.get(normalized_path)
        if cached and cached["mtime"] == mtime and cached["data"]:
            return jsonify({"data": {**cached["data"], "source": "excel-cache"}})

        parsed = parse_asan_shipping_excel(file_path)
        try:
            page_num = max(1, int(page or 1))
            page_limit = max(1, min(10000, int(page_size or 5000)))
            start = (page_num - 1) * page_limit
            end = start + page_limit
            all_rows = parsed.get("data") or []
            res = {
                **parsed,
                "data": all_rows[start:end],
                "file_modified_at": datetime.fromtimestamp(mtime, tz=KST).isoformat(),
                "source": "excel-cache",
                "total": len(all_rows),
                "page": page_num,
                "page_size": page_limit,
            }
            shipping_cache.pop(normalized_path, None)
            return jsonify({"data": res})
        finally:
            parsed = None
            gc.collect()
        
    except Exception as e:
        app.logger.error(f"선적관리 파싱 오류: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/vehicle-tracking/export/zip", methods=["GET"])
def export_zip_photos():
    """선택한 트립들의 사진 ZIP 압축 다운로드"""
    import zipfile
    ids_str = request.args.get("ids", "")
    if not ids_str: return "IDs missing", 400
    ids = ids_str.split(",")
    
    try:
        res = supabase.from_("vehicle_trips").select("id, vehicle_number, started_at, photos").in_("id", ids).execute()
        
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as zf:
            for trip in res.data:
                photos = trip.get("photos", [])
                vnum = trip.get("vehicle_number", "unknown")
                date = trip.get("started_at", "000000")[:10]
                for i, p in enumerate(photos):
                    key = p.get("key")
                    if key:
                        photo_url = f"{SUPABASE_URL}/storage/v1/object/public/vehicle-photos/{key}"
                        img_data = urlopen(photo_url).read()
                        zf.writestr(f"{date}_{vnum}_{i+1}.jpg", img_data)
        
        output.seek(0)
        return Response(output.read(), mimetype="application/zip", headers={"Content-Disposition": f"attachment; filename=photos.zip"})
    except Exception as e:
        return str(e), 500

@app.route("/api/debug/view", methods=["GET"])
def view_debug_log():
    """파일에 저장된 디버그 로그를 보여줍니다."""
    log_file = Path("debug_app.log")
    if not log_file.exists(): return "공개된 로그가 아직 없습니다."
    return send_file(log_file, mimetype="text/plain")

def _daemon_available():
    try:
        r = urlopen(Request(DAEMON_URL + "/health", method="GET"), timeout=2)
        return r.getcode() == 200
    except:
        return False

def run_runner(cmd, extra_args=None, env=None):
    args = [sys.executable, str(RUNNER), cmd] + (extra_args or [])
    env = dict(os.environ) if env is None else {**os.environ, **env}
    env["PYTHONIOENCODING"] = "utf-8"
    try:
        result = subprocess.run(args, cwd=str(ELSBOT_DIR), capture_output=True, text=True, encoding="utf-8", timeout=300, env=env)
        return result
    except Exception as e:
        app.logger.exception(f"Subprocess failed: {e}")
        raise

@app.route("/api/els/capabilities", methods=["GET"])
def capabilities():
    # 백엔드 내부에서 데몬 상태 확인 시 3회 재시도 로직 도입 (안정성 극대화)
    daemon_status = {
        "available": False,
        "driver_active": False,
        "user_id": None,
        "login_failures": 0,
        "max_login_attempts": 3,
        "login_protected": False,
    }
    
    for attempt in range(1, 4):
        try:
            r = urlopen(Request(DAEMON_URL + "/health", method="GET"), timeout=3)
            if r.getcode() == 200:
                raw = r.read().decode("utf-8")
                data = json.loads(raw)
                daemon_status["available"] = True
                daemon_status["driver_active"] = data.get("driver_active", False)
                daemon_status["user_id"] = data.get("user_id")
                daemon_status["login_failures"] = data.get("login_failures", 0)
                daemon_status["max_login_attempts"] = data.get("max_login_attempts", 3)
                daemon_status["login_protected"] = data.get("login_protected", False)
                
                # 살아있음이 확인되면 즉시 루프 탈출
                if daemon_status["driver_active"]:
                    break
        except Exception as e:
            app.logger.warning(f"[세션체크] {attempt}회차 시도 실패: {e}")
        
        if attempt < 3:
            time.sleep(0.5) # 잠시 대기 후 재시도
    
    # [v4.9.8] 좀비 잠금 해제: 전체 작업 5분 초과 시 강제 종료
    if global_progress.get("is_running") and global_progress.get("start_time"):
        if time.time() - global_progress["start_time"] > 300:
            app.logger.warning(f"[좀비복구] {global_progress.get('user')}의 작업이 5분을 초과하여 강제 종료 처리합니다.")
            global_progress["is_running"] = False
            global_progress["completed"] = global_progress["total"]

    # [v4.9.8] 유휴 잠금 해제: 마지막 개별 조회 완료 후 3분간 추가 조회 없으면 종료로 간주
    if global_progress.get("is_running") and global_last_activity_time > 0:
        idle = time.time() - global_last_activity_time
        if idle > 180:  # 3분 무활동
            app.logger.warning(f"[유휴복구] 마지막 조회 후 {int(idle)}초 무활동. 잠금 해제")
            global_progress["is_running"] = False
            global_progress["completed"] = global_progress["total"]

    return jsonify({
        "available": daemon_status["available"],
        "driver_active": daemon_status["driver_active"],
        "user_id": daemon_status.get("user_id"),
        "progress": global_progress,
        "workers": data.get("workers", []) if 'data' in locals() else [],
        "max_drivers": data.get("max_drivers", 3) if 'data' in locals() else 3,
        "login_failures": daemon_status["login_failures"],
        "max_login_attempts": daemon_status["max_login_attempts"],
        "login_protected": daemon_status["login_protected"],
        "parseAvailable": True
    })

@app.route("/api/els/config", methods=["GET"])
def config_get():
    if not CONFIG_PATH.exists():
        return jsonify({"hasSaved": False, "defaultUserId": ""})
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        uid = data.get("user_id", "")
        mtime = os.path.getmtime(str(CONFIG_PATH))
        dt = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({
            "hasSaved": bool(uid and data.get("user_pw")), 
            "defaultUserId": uid,
            "lastSaved": dt
        })
    except:
        return jsonify({"hasSaved": False, "defaultUserId": ""})

@app.route("/api/els/config", methods=["POST"])
def config_post():
    data = request.get_json(silent=True)
    if not data: return jsonify({"error": "No data"}), 400
    uid = (data.get("userId") or "").strip()
    pw = data.get("userPw") or ""
    if not uid or not pw: return jsonify({"error": "ID/PW required"}), 400
    
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(json.dumps({"user_id": uid, "user_pw": pw}, ensure_ascii=False, indent=2), encoding="utf-8")
    
    mtime = os.path.getmtime(str(CONFIG_PATH))
    dt = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
    
    return jsonify({"success": True, "defaultUserId": uid, "lastSaved": dt})

@app.route("/api/els/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    use_saved = data.get("useSavedCreds", True)
    uid = data.get("userId") or ""
    pw = data.get("userPw") or ""
    show_browser = data.get("showBrowser", False)

    if not _daemon_available():
        return jsonify({"ok": False, "error": "ELS 데몬이 실행되지 않았습니다."}), 404

    def generate():
        try:
            body = json.dumps({"useSavedCreds": use_saved, "userId": uid, "userPw": pw, "showBrowser": show_browser}, ensure_ascii=False).encode("utf-8")
            req = Request(DAEMON_URL + "/login", data=body, method="POST", headers={"Content-Type": "application/json"})
            
            login_result = {}
            def _thread_login():
                try:
                    # [NAS 최적화] 5개 세션 부팅(Stagger 15s)을 고려하여 400초로 연장
                    r = urlopen(req, timeout=400)
                    login_result['resp'] = json.loads(r.read().decode("utf-8"))
                except Exception as e:
                    login_result['error'] = str(e)

            t = threading.Thread(target=_thread_login)
            t.start()

            sent_logs = set()
            retry_count = 0
            while t.is_alive():
                try:
                    # 데몬의 /logs 엔드포인트에서 최신 로그 긁어오기
                    l_req = urlopen(DAEMON_URL + "/logs", timeout=5)
                    l_data = json.loads(l_req.read().decode("utf-8"))
                    has_new_daemon_log = False
                    for line in l_data.get("log", []):
                        if line not in sent_logs:
                            yield f"LOG:{line}\n"
                            sent_logs.add(line)
                            has_new_daemon_log = True
                    
                    if has_new_daemon_log:
                        retry_count = 0 # 데몬 로그가 찍히면 백엔드 심박 타이머 리셋 (사용자 혼란 방지)
                except: pass
                
                # 데몬으로부터 새로운 로그가 한동안 없을 때만 백엔드 심박 로그를 출력
                retry_count += 1
                if retry_count >= 30: 
                    yield f"LOG:[백엔드] 세션 초기화 대기 중... ({retry_count}s)\n"
                    # 심박 로그 출력 후 0으로 리셋하여 간격 유지
                    retry_count = 0 
                
                time.sleep(1)
            
            t.join()
            
            if 'error' in login_result:
                yield f"LOG:![오류] 데몬 로그인 실패: {login_result['error']}\n"
                yield "RESULT:" + json.dumps({"ok": False, "error": login_result['error']}) + "\n"
            else:
                final = login_result.get('resp', {})
                for line in final.get("log", []):
                    if line not in sent_logs:
                        yield f"LOG:{line}\n"
                
                if not final.get("ok") and final.get("error") == "LOGIN_ERROR_CREDENTIALS":
                     yield f"LOG:![경고] 이트랜스 계정 정보가 틀립니다.\n"
                
                yield "RESULT:" + json.dumps(final, ensure_ascii=False) + "\n"

        except Exception as e:
            yield f"LOG:![점검] 로그인 스트리밍 중 오류: {e}\n"

    return Response(generate(), mimetype="text/plain; charset=utf-8", headers={
        "X-Accel-Buffering": "no",
        "Cache-Control": "no-cache"
    })

    extra = []
    if not use_saved: extra.extend(["--user-id", uid, "--user-pw", pw])
    r = run_runner("login", extra_args=extra)
    out = (r.stdout or "").strip()
    try:
        start, end = out.rfind("{"), out.rfind("}")
        obj = json.loads(out[start : end + 1]) if start >= 0 else json.loads(out)
        return jsonify(obj)
    except:
        return jsonify({"ok": False, "error": "응답 파싱 실패"})

@app.route("/api/els/stop-daemon", methods=["POST"])
def stop_daemon():
    global global_progress
    if _daemon_available():
        try:
            req = Request(DAEMON_URL + "/stop", method="POST")
            urlopen(req, timeout=5)
            # [추가] 데몬 중지 시 백엔드의 진행률 락도 강제 해제
            global_progress["is_running"] = False
            global_progress["completed"] = global_progress.get("total", 0)
            return jsonify({"ok": True})
        except:
            pass
    return jsonify({"ok": False, "error": "데몬 중지 실패"})

def _stream_run_daemon(containers, use_saved, uid, pw, show_browser=False):
    from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
    global global_progress, global_last_activity_time
    
    final_rows = []
    headers = ["컨테이너번호", "No", "수출입", "구분", "터미널", "MOVE TIME", "모선", "항차", "선사", "적공", "SIZE", "POD", "POL", "차량번호", "RFID"]
    
    # 진행률 초기화 (시작 시간 기록하여 데드락 감지용 활용)
    global_progress = {
        "total": len(containers), 
        "completed": 0, 
        "is_running": True, 
        "start_time": time.time(),
        "user": uid
    }
    global_last_activity_time = time.time()
    yield f"LOG:병렬 조회를 시작합니다. (대상: {len(containers)}건, 병렬: 3개 세션 구동)\n"
    
    try:
        def fetch_container(cn):
            cn = cn.strip().upper()
            if not cn: return []
            st = time.time()
            try:
                body = json.dumps({"userId": uid, "userPw": pw, "containerNo": cn, "showBrowser": show_browser}, ensure_ascii=False).encode("utf-8")
                req = Request(DAEMON_URL + "/run", data=body, method="POST", headers={"Content-Type": "application/json"})
                resp = urlopen(req, timeout=120)
                res_json = json.loads(resp.read().decode("utf-8"))
                worker_id = res_json.get("worker_id", "?")
                daemon_id = res_json.get("daemon_id", "1") # [추가] 데몬 식별 ID 수집
                return res_json.get("result", []), cn, res_json.get("error"), round(time.time() - st, 1), worker_id, daemon_id
            except Exception as e:
                return [[cn, "ERROR", str(e)] + [""]*12], cn, str(e), round(time.time() - st, 1), "?", "ER"

        sent_daemon_logs = set()
        with ThreadPoolExecutor(max_workers=3) as executor:
            # 딕셔너리에 (컨테이너번호, 재시도횟수) 상태 저장
            futures = {executor.submit(fetch_container, cn): (cn, 0) for cn in containers}
            
            log_poll_counter = 0
            while futures:
                # 🎯 [실시간 폴링 지연] 매 루프마다 가져오지 않고 3회에 한 번만(약 1.5초 주기) 가져옴으로써 NAS 부하 감소
                log_poll_counter += 1
                if log_poll_counter >= 3:
                    try:
                        l_req = urlopen(DAEMON_URL + "/logs", timeout=1)
                        l_data = json.loads(l_req.read().decode("utf-8"))
                        for line in l_data.get("log", []):
                            if line not in sent_daemon_logs:
                                yield f"LOG:{line}\n"
                                sent_daemon_logs.add(line)
                    except: pass
                    log_poll_counter = 0

                # 완료된 작업이 있는지 체크 (0.5초 대기)
                done, not_done = wait(futures.keys(), timeout=0.5, return_when=FIRST_COMPLETED)
                for f in done:
                    cn, retries = futures[f]
                    rows, _, err, elapsed, worker_id, daemon_id = f.result()
                    
                    # 🎯 에러 발생 시 재시도 로직 (1회 한정하여 다른 데몬에게 하청/재조회 시도)
                    if err and retries < 1:
                        yield f"LOG:⚠️ [D#{daemon_id}-B#{worker_id}] [{cn}] 오류 발생, 재조회 시도 중... ({err})\n"
                        new_f = executor.submit(fetch_container, cn)
                        futures[new_f] = (cn, retries + 1)
                        del futures[f]
                        continue
                    
                    final_rows.extend(rows)
                    global_progress["completed"] += 1
                    global_last_activity_time = time.time()  # [v4.9.8] 개별 조회 완료 시마다 갱신
                    
                    # 안전한 JSON 전송을 위한 처리
                    def _safe_val(v):
                        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
                        return v
                    partial_rows = [[_safe_val(cell) for cell in row] for row in rows]

                    if err:
                        yield f"LOG:❌ [D#{daemon_id}-B#{worker_id}] [{global_progress['completed']}/{global_progress['total']}] [{cn}] 실패 ({elapsed}s): {err}\n"
                        # 에러 상태도 즉시 프론트에 전달
                        yield "RESULT_PARTIAL:" + json.dumps({"result": partial_rows}, ensure_ascii=False) + "\n"
                    else:
                        yield f"LOG:✔ [D#{daemon_id}-B#{worker_id}] [{global_progress['completed']}/{global_progress['total']}] [{cn}] 완료 ({len(rows)}건) ({elapsed}s)\n"
                        # 성공 결과 전달 (내역 없음 포함)
                        yield "RESULT_PARTIAL:" + json.dumps({"result": partial_rows}, ensure_ascii=False) + "\n"
                    
                    del futures[f]
    finally:
        global_progress["is_running"] = False
        global_progress["completed"] = global_progress["total"]
    # Generate Final Result and Excel (이후 로직 동일)
    if final_rows:
        try:
            yield "LOG:조회 완료! 데이터 정리 중...\n"
            
            # 1. 데이터프레임 생성 및 정제 (가장 중요)
            df_all = pd.DataFrame(final_rows, columns=headers)
            df_clean = df_all.where(pd.notnull(df_all), None)
            
            # 안전하게 JSON용 리스트 변환
            def _safe_val(v):
                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
                return v
            rows_list = [[_safe_val(cell) for cell in row] for row in df_clean.values.tolist()]

            # 2. 화면에 데이터 먼저 뿌려주기 (이걸 먼저 해야 형이 바로 볼 수 있어!)
            token = str(uuid.uuid4()).replace("-", "")[:16]
            filename = f"els_result_{datetime.now().strftime('%y%m%d_%H%M%S')}.xlsx"
            
            result_obj = {
                "ok": True,
                "result": rows_list,
                "downloadToken": token, # 일단 토큰 발행
                "fileName": filename
            }
            yield "RESULT:" + json.dumps(result_obj, ensure_ascii=False) + "\n"
            yield "LOG:✔ 화면 데이터 전송 완료\n"

            # 3. 그 다음 엑셀 파일 생성 (실패해도 데이터는 보임)
            try:
                yield "LOG:엑셀 파일 생성 중...\n"
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter

                output = io.BytesIO()

                # 시트1: 최신현황 (No=1인 것만), 시트2: 전체이력
                df_latest = df_all[df_all['No'].astype(str) == '1'].drop_duplicates()
                df_full = df_all.drop_duplicates()

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_latest.to_excel(writer, index=False, sheet_name='최신현황')
                    df_full.to_excel(writer, index=False, sheet_name='전체이력')

                    # 서식 정의
                    header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
                    banip_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
                    suip_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
                    header_font = Font(bold=True)

                    for sheet_name in ['최신현황', '전체이력']:
                        ws = writer.sheets[sheet_name]

                        # (1) 제목행 서식: 옅은 파랑 배경 + 볼드
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font

                        # (2) 틀 고정 (1행)
                        ws.freeze_panes = ws['A2']

                        # (3) 자동 필터 (정렬 가능)
                        ws.auto_filter.ref = ws.dimensions

                        # (4) 컬럼 너비 자동 조절 (가장 긴 글씨 기준, 한글 폭 보정)
                        for col_idx, col_cells in enumerate(ws.columns, 1):
                            max_length = 0
                            for cell in col_cells:
                                try:
                                    val_str = str(cell.value) if cell.value is not None else ""
                                    cell_len = len(val_str)
                                    korean_extra = sum(1 for c in val_str if ord(c) > 127)
                                    cell_len += korean_extra
                                    if cell_len > max_length:
                                        max_length = cell_len
                                except:
                                    pass
                            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 3, 8)

                        # (5) 조건부 셀 색상: "반입" → 옅은 파랑, "수입" → 옅은 붉은색
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for cell in row:
                                val = str(cell.value or '')
                                if '반입' in val:
                                    cell.fill = banip_fill
                                elif '수입' in val:
                                    cell.fill = suip_fill

                output.seek(0)
                file_store[token] = output.read()
                yield "LOG:✔ 엑셀 생성 완료 (다운로드 가능)\n"
            except Exception as e:
                yield f"LOG:![알림] 엑셀 파일 생성은 실패했습니다 (데이터는 정상). 에러: {e}\n"
                # 엑셀 실패 시 토큰 무효화
                if token in file_store: del file_store[token]

            yield "LOG:----------------------------------------\n"
            yield "LOG:모든 작업이 완료되었습니다. 결과창을 확인하세요!\n"
                
        except Exception as e:
             yield f"LOG:[치명적에러] 최종 집계 중 오류 발생: {e}\n"
             import traceback
             print(traceback.format_exc())
    else:
        yield "LOG:조회 결과 데이터가 하나도 없습니다.\n"

def _stream_run(containers, use_saved, uid, pw, show_browser=False):
    """데몬 모드로 컨테이너 조회를 스트리밍 실행한다."""
    if not _daemon_available():
        yield "LOG:[오류] ELS 데몬이 실행되지 않았습니다. 로그인을 먼저 수행하세요.\n"
        return
    yield from _stream_run_daemon(containers, use_saved, uid, pw, show_browser=show_browser)

@app.route("/api/els/run", methods=["POST"])
def run():
    data = request.get_json(silent=True) or {}
    return Response(_stream_run(data.get("containers", []), data.get("useSavedCreds", True), data.get("userId", ""), data.get("userPw", ""), data.get("showBrowser", False) ), mimetype="text/plain; charset=utf-8", headers={
        "X-Accel-Buffering": "no",
        "Cache-Control": "no-cache"
    })

@app.route("/api/els/parse-xlsx", methods=["POST"])
def parse_xlsx():
    f = request.files["file"]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        f.save(tmp.name)
        r = run_runner("parse", extra_args=[tmp.name])
        os.unlink(tmp.name)
    return jsonify(json.loads(r.stdout))

@app.route("/api/els/download/<token>", methods=["GET"])
def download(token):
    buf = file_store.pop(token, None)
    if not buf: return "Expired", 404
    name = request.args.get("filename") or "els_result.xlsx"
    return Response(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{name}"})

@app.route("/api/nas/files", methods=["POST"])
def upload_nas_file():
    """공지사항 등 첨부파일 업로드 (나스 저장)"""
    if "file" not in request.files: return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    rel_path = request.form.get("path", "/uploads")
    
    # 나스 내 저장 경로 설정
    save_dir = Path("/app/data") / rel_path.strip("/")
    save_dir.mkdir(parents=True, exist_ok=True)
    
    # 파일명 클리닝 및 중복 방지
    fname = secure_filename(f.filename)
    save_path = save_dir / fname
    
    f.save(str(save_path))
    return jsonify({"success": True, "path": str(rel_path.strip("/") + "/" + fname)})

@app.route("/api/nas/files", methods=["GET"])
def download_nas_file():
    """나스에 저장된 첨부파일 다운로드"""
    rel_path = request.args.get("path")
    if not rel_path: return "Path missing", 400
    
    full_path = Path("/app/data") / rel_path.strip("/")
    if not full_path.exists(): return "File not found", 404
    
    name = request.args.get("name") or full_path.name
    return send_file(str(full_path), as_attachment=True, download_name=name)

@app.route("/api/els/logout", methods=["POST"])
def logout():
    if _daemon_available():
        req = Request(DAEMON_URL + "/quit", data=b"{}", method="POST", headers={"Content-Type": "application/json"})
        try: urlopen(req, timeout=5)
        except: pass
    return jsonify({"ok": True})

@app.route("/api/els/template", methods=["GET"])
def template():
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["컨테이너넘버"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="template.xlsx")

@app.route("/api/last-result", methods=["GET"])
def get_last_result():
    if LAST_RESULT_FILE.exists():
        try:
            with open(LAST_RESULT_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return jsonify({"ok": True, **data})
        except: pass
    return jsonify({"ok": False, "message": "최근 데이터가 없습니다."})

@app.route("/api/els/screenshot", methods=["GET"])
def screenshot():
    """데몬으로부터 스크린샷 이미지를 가져와 릴레이합니다. (디버깅용)"""
    if _daemon_available():
        try:
            r = urlopen(DAEMON_URL + "/screenshot", timeout=5)
            return Response(r.read(), mimetype='image/png')
        except HTTPError as e:
            # 데몬이 404 등을 반환한 경우 (스크린샷이 아직 없음 등)
            error_msg = f"Daemon returned {e.code}"
            try: error_msg += " - " + e.read().decode('utf-8')
            except: pass
            return jsonify({"error": error_msg}), e.code
        except URLError as e:
            # 데몬 연결 자체가 안되는 경우
            return jsonify({"error": f"Daemon connection failed: {e.reason}"}), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    return jsonify({"error": "Daemon not available"}), 404

from web_vectorizer import process_web_attachments, init_supabase as init_web_supabase
import asyncio

@app.route('/api/vectorize/nas', methods=['POST'])
def trigger_nas_vectorize():
    return jsonify({"status": "disabled", "message": "NAS vectorization is disabled due to system load."}), 400

@app.route('/api/vectorize/web', methods=['POST'])
def trigger_web_vectorize():
    """Trigger Web Attachment vectorization (Phase 5 Extension)."""
    if not supabase:
        return jsonify({"error": "Supabase client not initialized"}), 500
    
    try:
        init_web_supabase(supabase)
        success = process_web_attachments()
        return jsonify({"ok": success})
    except Exception as e:
        app.logger.error(f"Web vectorize error: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 2929))
    app.logger.info(f"Backend Server Ready with CORS on port {port}")
    app.run(host="0.0.0.0", port=port, threaded=True)
