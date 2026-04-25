import os
import hashlib
import logging
from pathlib import Path
from datetime import datetime, timezone
import time
import openpyxl
import fitz  # PyMuPDF
from docx import Document
import pytesseract
from PIL import Image
import textract
import gc
import requests
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter

logger = logging.getLogger(__name__)

# chunking params
MAX_CHUNK_SIZE = 800
CHUNK_OVERLAP = 150

# Supported extensions mapping to type
SUPPORTED_EXTS = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".doc": "doc",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
    ".txt": "txt",
    ".hwpx": "hwpx",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".gif": "image",
}

def supabase_retry_execute(query_func, max_retries=5, base_delay=2):
    """Supabase 호출 실패 시(특히 429 Rate Limit) 재시도 로직"""
    for attempt in range(max_retries):
        try:
            return query_func().execute()
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "Too Many Requests" in err_str:
                delay = base_delay * (2 ** attempt)
                logger.warning(f"⚠️ Supabase Rate Limit 감지 (시도 {attempt+1}/{max_retries}). {delay}초 후 재시도... 에러: {err_str}")
                time.sleep(delay)
            else:
                # 일반적인 에러는 바로 발생시킴
                raise e
    return query_func().execute() # 마지막 시도

def get_file_hash(filepath):
    """파일의 MD5 해시를 계산하여 변경 여부 감지"""
    hash_md5 = hashlib.md5()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
    except Exception as e:
        logger.error(f"Hash calculation failed for {filepath}: {e}")
        return "error"
    return hash_md5.hexdigest()

def extract_text_pypdf(filepath):
    text = ""
    doc = None
    try:
        doc = fitz.open(filepath)
        for page in doc:
            text += page.get_text("text") + "\n\n"
    except Exception as e:
        logger.error(f"PDF extraction failed for {filepath}: {e}")
    finally:
        if doc:
            doc.close()
    return text

def extract_text_docx(filepath):
    text = ""
    try:
        doc = Document(filepath)
        for para in doc.paragraphs:
            if para.text.strip():
                text += para.text + "\n"
    except Exception as e:
        logger.error(f"DOCX extraction failed for {filepath}: {e}")
    return text

def extract_text_hwpx(filepath):
    """HWPX 파일(XML 기반 ZIP)에서 텍스트 추출"""
    import zipfile
    import xml.etree.ElementTree as ET
    text = ""
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            # Contents/ 폴더 내의 section*.xml 파일들이 실제 본문 내용을 담고 있음
            xml_files = [f for f in z.namelist() if 'Contents/section' in f and f.endswith('.xml')]
            xml_files.sort() # 순서대로 합침
            for xml_file in xml_files:
                with z.open(xml_file) as f:
                    root = ET.fromstring(f.read())
                    # 모든 텍스트 노드 추출 (태그 상관없이 텍스트만)
                    for node in root.iter():
                        if node.text and node.text.strip():
                            text += node.text.strip() + " "
                    text += "\n"
    except Exception as e:
        logger.error(f"HWPX extraction failed for {filepath}: {e}")
    return text

def extract_sheets_xlsx(filepath):
    """엑셀 파일을 시트별로 분리하여 텍스트와 해시 추출"""
    sheets_data = []
    wb = None
    try:
        # read_only=True는 메모리 사용량을 획기적으로 줄여줍니다.
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = f"--- 시트명: {sheet_name} ---\n"
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if row_vals:
                    sheet_text += " | ".join(row_vals) + "\n"
                    row_count += 1
            
            if row_count > 0:
                sheet_hash = hashlib.md5(sheet_text.encode('utf-8')).hexdigest()
                sheets_data.append({
                    "name": sheet_name,
                    "text": sheet_text,
                    "hash": sheet_hash
                })
    except Exception as e:
        logger.error(f"XLSX extraction failed for {filepath}: {e}")
    finally:
        if wb:
            wb.close()
    return sheets_data

def chunk_text(text, max_len=MAX_CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    chunks = []
    start = 0
    while start < len(text):
        end = start + max_len
        chunk = text[start:end]
        chunks.append(chunk)
        start += (max_len - overlap)
    return chunks

def process_nas_directory(supabase, raw_dir, branch_name="NAS자료"):
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        logger.error("GEMINI_API_KEY is not set.")
        return {"error": "GEMINI_API_KEY is missing"}

    target_dir = Path(raw_dir)
    if not target_dir.exists() or not target_dir.is_dir():
        logger.error(f"Target directory {target_dir} does not exist.")
        return {"error": f"Dir {target_dir} not found"}

    processed = 0
    skipped = 0
    error_cnt = 0
    
    # 1. 대상 파일 리스트업
    all_target_files = []
    logger.info(f"🔍 [{branch_name}] ({target_dir}) 파일 탐색 중...")
    for root, dirs, files in os.walk(str(target_dir)):
        skip_words = ["#recycle", "@eaDir", "보안", "RESTRICTED", "PRIVATE", "E_임직원캐비넷", "F_Backup", "G_Downloads"]
        if any(word in root for word in skip_words):
            continue
            
        now_ts = time.time()
        one_week_ago = now_ts - (7 * 24 * 60 * 60)
            
        for file in files:
            # [v5.8.4] 임시 파일 및 캐시 파일 제외 (~$, .tmp 등)
            if file.startswith("~$") or ".tmp" in file.lower() or file.startswith("."):
                continue
                
            filepath = Path(root) / file
            if filepath.suffix.lower() in SUPPORTED_EXTS:
                try:
                    # 제한 용량을 20MB로 하향 조정 (나스 메모리 보호)
                    # [v5.8.1] 7일 이내 수정된 파일만 파싱 (사용자 정책 반영)
                    # 단, 기존에 인덱싱된 적이 없는 파일은 최초 1회 파싱을 시도할 수도 있으나, 
                    # 12k 스캔 부하 방지를 위해 엄격하게 mtime 7일 이내로 제한함.
                    mtime = filepath.stat().st_mtime
                    if mtime >= one_week_ago:
                        if filepath.stat().st_size <= 20 * 1024 * 1024:
                            all_target_files.append(filepath)
                except:
                    pass
    
    total_files = len(all_target_files)
    logger.info(f"✅ [{branch_name}] 총 {total_files}개의 대상 파일을 찾았습니다.")

    # 2. 기존 인덱스 일괄 조회 (병목 해결 핵심)
    logger.info(f"📂 [{branch_name}] 기존 DB 인덱스 일괄 조회 중...")
    try:
        existing_data = supabase.table("nas_file_index").select("id, path, content_hash, is_indexed").eq("branch", branch_name).execute().data
        index_map = {item['path']: item for item in existing_data}
        logger.info(f"✅ 기존 인덱스 {len(index_map)}개 로드 완료.")
    except Exception as e:
        logger.error(f"❌ 기존 인덱스 조회 실패: {e}")
        index_map = {}

    # 3. 파일별 처리 루프
    for i, filepath in enumerate(all_target_files):
        filename = filepath.name
        ext = filepath.suffix.lower()
        file_path_str = str(filepath.resolve())
        progress_str = f"📄 [{i+1}/{total_files}]"

        try:
            # 해시 계산 및 중복 체크
            current_hash = get_file_hash(file_path_str)
            existing_item = index_map.get(file_path_str)
            
            if existing_item and existing_item.get("is_indexed") and existing_item.get("content_hash") == current_hash:
                skipped += 1
                if skipped % 50 == 0: # 50개마다 하나씩만 로그 찍어서 가독성 유지
                    logger.info(f"⏩ {progress_str} [SKIP] 이전 작업 완료된 파일들 통과 중...")
                continue

            # [v5.5.1] 진행 상황 가독성 강화
            logger.info(f"⚙️ {progress_str} 처리 시작: {filename} ({ext})")
            
            # 텍스트 추출 및 처리 (엑셀과 일반 파일 분리)
            if ext in [".xlsx", ".xlsm"]:
                sheets = extract_sheets_xlsx(file_path_str)
                if not sheets:
                    skipped += 1
                    continue
                
                # 시트별 증분 파싱 로직 (v5.8.2)
                # 기존 해당 파일의 모든 청크의 메타데이터를 가져와서 시트별 해시 비교
                existing_chunks_res = supabase.table("document_chunks").select("metadata").eq("source_id", file_path_str).execute()
                existing_sheet_map = {}
                for c in (existing_chunks_res.data or []):
                    m = c.get("metadata", {})
                    s_n = m.get("sheet_name")
                    s_h = m.get("sheet_hash")
                    if s_n: existing_sheet_map[s_n] = s_h

                found_sheet_names = [s["name"] for s in sheets]
                # 엑셀에서 삭제된 시트의 청크 정리
                for ex_name in list(existing_sheet_map.keys()):
                    if ex_name not in found_sheet_names:
                        logger.info(f"🗑️ 시트 삭제 감지: {ex_name} (파일: {filename})")
                        supabase.table("document_chunks").delete().eq("source_id", file_path_str).filter("metadata->>sheet_name", "eq", ex_name).execute()

                total_new_chunks = 0
                for s in sheets:
                    s_name = s["name"]
                    s_hash = s["hash"]
                    s_text = s["text"]

                    if s_name in existing_sheet_map and existing_sheet_map[s_name] == s_hash:
                        # logger.info(f"⏩ 시트 변경 없음 스킵: {s_name}")
                        continue
                    
                    logger.info(f"⚙️ 시트 파싱 시작: {s_name} (해시 변경 또는 신규)")
                    # 기존 시트 청크 삭제
                    supabase.table("document_chunks").delete().eq("source_id", file_path_str).filter("metadata->>sheet_name", "eq", s_name).execute()
                    
                    chunks = chunk_text(s_text)
                    chunk_batch = []
                    for idx, chunk in enumerate(chunks):
                        if not chunk.strip(): continue
                        
                        time.sleep(0.4) # Rate Limit 방어
                        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-embedding-001:embedContent?key={api_key}"
                        payload = {
                            "model": "models/gemini-embedding-001",
                            "content": {"parts": [{"text": chunk}]},
                            "outputDimensionality": 768
                        }
                        
                        try:
                            resp = requests.post(url, json=payload, timeout=10)
                            if resp.status_code == 200:
                                embedding = resp.json()['embedding']['values']
                                chunk_batch.append({
                                    "source_type": "nas_file",
                                    "source_id": file_path_str,
                                    "source_version": datetime.now().strftime('%Y-%m-%d'),
                                    "chunk_index": idx,
                                    "content": chunk,
                                    "metadata": {
                                        "filename": filename, 
                                        "branch": branch_name, 
                                        "extension": ext,
                                        "sheet_name": s_name,
                                        "sheet_hash": s_hash
                                    },
                                    "embedding": embedding
                                })
                            
                            if len(chunk_batch) >= 10:
                                supabase_retry_execute(lambda: supabase.table("document_chunks").insert(chunk_batch))
                                total_new_chunks += len(chunk_batch)
                                chunk_batch = []
                                time.sleep(0.2)
                        except Exception as e:
                            logger.error(f"Embedding error at {s_name} chunk {idx}: {e}")

                    if chunk_batch:
                        supabase_retry_execute(lambda: supabase.table("document_chunks").insert(chunk_batch))
                        total_new_chunks += len(chunk_batch)
                
                # 인덱스 업데이트
                supabase_retry_execute(lambda: supabase.table("nas_file_index").upsert({
                    "path": file_path_str, "filename": filename, "extension": ext, "branch": branch_name,
                    "content_hash": current_hash, "is_indexed": True, "updated_at": datetime.now(timezone.utc).isoformat()
                }))
                processed += 1
                logger.info(f"✅ {filename} 처리 완료 (새로운 청크: {total_new_chunks}개)")

            else:
                # 일반 파일 (PDF, DOCX 등) 처리 로직
                text = ""
                if ext == ".pdf":
                    text = extract_text_pypdf(file_path_str)
                elif ext == ".docx":
                    text = extract_text_docx(file_path_str)
                elif ext == ".txt":
                    try:
                        with open(file_path_str, "r", encoding="utf-8") as f: text = f.read()
                    except:
                        with open(file_path_str, "r", encoding="euc-kr") as f: text = f.read()
                elif ext in [".png", ".jpg", ".jpeg", ".gif"]:
                    image = None
                    try:
                        image = Image.open(file_path_str)
                        text = pytesseract.image_to_string(image, lang='kor+eng')
                    finally:
                        if image: image.close()
                elif ext == ".doc":
                    try:
                        text = textract.process(file_path_str).decode('utf-8', errors='ignore')
                    except:
                        text = ""
                elif ext == ".hwpx":
                    text = extract_text_hwpx(file_path_str)

                if not text or len(text.strip()) < 10:
                    logger.warning(f"⚠️ {progress_str} [SKIP] {filename}: 텍스트 부족")
                    skipped += 1
                    continue

                # 청킹 및 임베딩 (기존 로직 유지)
                chunks = chunk_text(text)
                chunk_batch = []
                supabase_retry_execute(lambda: supabase.table("document_chunks").delete().eq("source_id", file_path_str))

                for idx, chunk in enumerate(chunks):
                    if not chunk.strip(): continue
                    time.sleep(0.4)
                    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-embedding-001:embedContent?key={api_key}"
                    payload = {"model": "models/gemini-embedding-001", "content": {"parts": [{"text": chunk}]}, "outputDimensionality": 768}
                    try:
                        resp = requests.post(url, json=payload, timeout=10)
                        if resp.status_code == 200:
                            embedding = resp.json()['embedding']['values']
                            chunk_batch.append({
                                "source_type": "nas_file", "source_id": file_path_str, "source_version": datetime.now().strftime('%Y-%m-%d'),
                                "chunk_index": idx, "content": chunk, "embedding": embedding,
                                "metadata": {"filename": filename, "branch": branch_name, "extension": ext}
                            })
                        if len(chunk_batch) >= 10:
                            supabase_retry_execute(lambda: supabase.table("document_chunks").insert(chunk_batch))
                            chunk_batch = []
                            time.sleep(0.2)
                    except Exception as e:
                        logger.error(f"Embedding error at {filename} chunk {idx}: {e}")

                if chunk_batch:
                    supabase_retry_execute(lambda: supabase.table("document_chunks").insert(chunk_batch))

                supabase_retry_execute(lambda: supabase.table("nas_file_index").upsert({
                    "path": file_path_str, "filename": filename, "extension": ext, "branch": branch_name,
                    "content_hash": current_hash, "is_indexed": True, "chunk_count": len(chunks),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }))
                processed += 1
            
            # [v5.5.1] NAS 보호를 위한 강제 휴식 및 메모리 정리
            time.sleep(2)  # 파일 하나당 2초 휴식 (CPU 과부하 방지)
            gc.collect()   # 메모리 즉시 회수
            
        except Exception as e:
            logger.error(f"❌ Error processing {filename}: {e}")
            error_cnt += 1
        finally:
            gc.collect()

    logger.info(f"🎉 [{branch_name}] 작업 완료! (처리: {processed}, 스킵: {skipped}, 에러: {error_cnt})")
    return {"processed": processed, "skipped": skipped, "errors": error_cnt}
