"""
웹 연동용 러너. els_bot.py 핵심 로직(함수)은 그대로 사용하고, CLI/JSON 입출력만 담당.
- parse: xlsx 경로 → stdout JSON {"containers": [...]}
- run: containers + creds → stdout JSON {"log": [...], "sheet1": [...], "sheet2": [...], "output_path": "..."}
"""
import sys
import os
import json
import time
import datetime
import tempfile
import argparse

# els_bot 핵심 변경 없이 기존 함수만 import
from els_bot import (
    load_config,
    login_and_prepare,
    solve_input_and_search,
    scrape_hyper_verify,
)

def parse_xlsx(path):
    """A1=컨테이너넘버(제목), A2부터 컨테이너 번호."""
    import pandas as pd
    if not os.path.isfile(path):
        return []
    df = pd.read_excel(path)
    if df.empty or len(df) < 2:
        return []
    col0 = df.iloc[1:, 0].dropna().astype(str).str.strip().str.upper()
    return col0.tolist()


def run_search(containers, user_id=None, user_pw=None, driver=None, keep_alive=False, log_callback=None):
    """
    containers: list of container numbers (str).
    user_id, user_pw: None이면 load_config() 사용.
    driver: 기존 WebDriver가 있으면 재사용(로그인 생략).
    keep_alive: True면 조회 후 driver.quit() 하지 않음(세션 유지).
    log_callback: 호출 가능 객체(msg)면 각 로그마다 호출(실시간 스트리밍용).
    returns: (log_lines, sheet1_data, sheet2_data, output_excel_path or None)
    """
    config = load_config()
    u_id = user_id if user_id else config.get("user_id", "")
    u_pw = user_pw if user_pw else config.get("user_pw", "")
    if not driver and (not u_id or not u_pw):
        return (["[오류] 아이디/비밀번호가 없습니다."], [], [], None)

    headers = ["조회번호", "No", "수출입", "구분", "터미널", "MOVE TIME", "모선", "항차", "선사", "적공", "SIZE", "POD", "POL", "차량번호", "RFID"]
    log_lines = []
    final_rows = []

    def log(msg):
        log_lines.append(msg)
        print(msg, flush=True)
        if log_callback:
            try:
                log_callback(msg)
            except Exception:
                pass

    own_driver = False
    if driver is None:
        try:
            log("엔진 예열 및 로그인 중...")
            driver = login_and_prepare(u_id, u_pw)
            if not driver:
                log("로그인 실패!")
                return (log_lines, [], [], None)
            log("로그인 성공.")
            own_driver = True
        except Exception as e:
            log(f"[예외] {e}")
            return (log_lines, [], [], None)
    else:
        log("기존 세션으로 조회 진행.")
    try:
        for cn_raw in containers:
            cn = str(cn_raw).strip().upper()
            if not cn:
                continue
            unit_start = time.time()
            log(f"[{cn}] 분석 중...")

            status = solve_input_and_search(driver, cn)
            if isinstance(status, str) and "오류" in status:
                dur = time.time() - unit_start
                log(f"  패스 ({status}) [{dur:.2f}s]")
                final_rows.append([cn, "ERROR", status] + [""] * 12)
                continue

            grid_text = None
            for _ in range(120):
                grid_text = scrape_hyper_verify(driver, cn)
                if grid_text:
                    break
                time.sleep(0.05)

            dur = time.time() - unit_start
            if grid_text:
                for line in grid_text.split("\n"):
                    final_rows.append([cn] + line.split("|"))
                log(f"  성공! [{dur:.2f}s]")
            else:
                log(f"  불일치/내역없음 [{dur:.2f}s]")
                final_rows.append([cn, "NODATA", "데이터 없음"] + [""] * 12)

        if not final_rows:
            return (log_lines, [], [], None)

        import pandas as pd
        from openpyxl.styles import PatternFill

        df_out = pd.DataFrame(final_rows)
        df_out.columns = headers[: df_out.shape[1]]
        sheet1_df = df_out[df_out.iloc[:, 1].astype(str) == "1"]
        now = datetime.datetime.now().strftime("%m%d_%H%M")
        fd, output_path = tempfile.mkstemp(suffix=f"_els_hyper_{now}.xlsx", prefix="els_")
        os.close(fd)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            sheet1_df.to_excel(writer, sheet_name="Sheet1", index=False)
            df_out.to_excel(writer, sheet_name="Sheet2", index=False)
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            blue = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            err = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for sn in ["Sheet1", "Sheet2"]:
                ws = writer.sheets[sn]
                for r in ws.iter_rows(min_row=2):
                    if len(r) > 2 and r[2].value == "수입":
                        r[2].fill = red
                    if len(r) > 3 and r[3].value == "반입":
                        r[3].fill = blue
                    if len(r) > 1 and str(r[1].value) in ["ERROR", "NODATA"]:
                        r[1].fill = err

        log(f"파일 생성 완료: {output_path}")

        def to_python(x):
            if hasattr(x, "item"):
                return x.item()
            if x != x:  # NaN
                return None
            return x

        sheet1_data = [[to_python(x) for x in row] for row in sheet1_df.values.tolist()]
        sheet2_data = [[to_python(x) for x in row] for row in df_out.values.tolist()]
        return (log_lines, sheet1_data, sheet2_data, output_path)

    except Exception as e:
        log_lines.append(f"[예외] {e}")
        print(f"[예외] {e}", flush=True)
        return (log_lines, [], [], None)
    finally:
        if driver and own_driver and not keep_alive:
            try:
                driver.quit()
            except Exception:
                pass


def login_only(user_id=None, user_pw=None):
    """로그인만 수행. 성공 시 (True, log_lines), 실패 시 (False, log_lines)."""
    config = load_config()
    u_id = user_id if user_id else config.get("user_id", "")
    u_pw = user_pw if user_pw else config.get("user_pw", "")
    if not u_id or not u_pw:
        return (False, ["[오류] 아이디/비밀번호가 없습니다."])
    driver = None
    try:
        driver = login_and_prepare(u_id, u_pw)
        if driver:
            return (True, ["로그인 성공.", "조회 페이지 대기 중."])
        return (False, ["로그인 실패!"])
    except Exception as e:
        return (False, [f"[예외] {e}"])
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd", required=True)
    p_parse = sub.add_parser("parse")
    p_parse.add_argument("path", help="container_list.xlsx path")
    p_login = sub.add_parser("login")
    p_login.add_argument("--user-id", default=None, help="ELS user id (optional, use config if omitted)")
    p_login.add_argument("--user-pw", default=None, help="ELS password (optional)")
    p_run = sub.add_parser("run")
    p_run.add_argument("--containers", required=True, help="JSON array of container numbers")
    p_run.add_argument("--user-id", default=None, help="ELS user id (optional, use config if omitted)")
    p_run.add_argument("--user-pw", default=None, help="ELS password (optional)")

    args = parser.parse_args()

    if args.cmd == "parse":
        containers = parse_xlsx(args.path)
        print(json.dumps({"containers": containers}, ensure_ascii=False))

    elif args.cmd == "login":
        ok, log_lines = login_only(user_id=args.user_id, user_pw=args.user_pw)
        out = {"ok": ok, "log": log_lines}
        if not ok:
            out["error"] = log_lines[-1] if log_lines else "Unknown error"
        print(json.dumps(out, ensure_ascii=False))

    elif args.cmd == "run":
        try:
            containers = json.loads(args.containers)
        except json.JSONDecodeError:
            print(json.dumps({"error": "Invalid JSON containers"}))
            sys.exit(1)
        if not containers:
            print(json.dumps({"error": "containers empty"}))
            sys.exit(1)
        log_lines, sheet1, sheet2, output_path = run_search(
            containers, user_id=args.user_id, user_pw=args.user_pw
        )
        out = {
            "log": log_lines,
            "sheet1": sheet1 or [],
            "sheet2": sheet2 or [],
            "output_path": output_path,
        }
        if "error" in str(log_lines):
            out["error"] = log_lines[-1] if log_lines else "Unknown error"
        print(json.dumps(out, ensure_ascii=False))


if __name__ == "__main__":
    main()
